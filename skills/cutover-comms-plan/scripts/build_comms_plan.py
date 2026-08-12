#!/usr/bin/env python3
"""Build a cutover communications plan workbook from a JSON spec.

Two modes:

  1. Default   - generates a formatted workbook from scratch (Comms Plan /
                 Plan Info / Reference sheets, dropdowns, filters).
  2. Template  - populates a copy of an existing .xlsx template, matching the
                 spec's fields to whatever headers the template already uses
                 and preserving its formatting.

Usage:
    python3 build_comms_plan.py --spec spec.json --out plan.xlsx
    python3 build_comms_plan.py --spec spec.json --out plan.xlsx \
        --template client_template.xlsx [--sheet "Comms Plan"] [--header-row 3]
    python3 build_comms_plan.py --list-fields

Requires openpyxl (pip install openpyxl).
"""

import argparse
import json
import re
import sys
from copy import copy
from datetime import date, datetime, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")


# --------------------------------------------------------------------------
# Canonical schema
# --------------------------------------------------------------------------
# (spec key, output header, column width, synonyms used to match template headers)
FIELDS = [
    ("id", "Comms ID", 10,
     ["id", "ref", "no", "num", "number", "item", "comm id", "comms id", "message id",
      "comms", "comms #", "comms no", "task #", "task no"]),
    ("milestone", "Cutover Milestone", 18,
     ["milestone", "timing", "phase", "when", "cutover milestone", "trigger",
      "t-minus", "stage", "sequence", "activity category", "comms category",
      "category", "comms type"]),
    ("send_date", "Planned Send Date", 16,
     ["date", "send date", "planned date", "planned send date", "issue date",
      "target date", "distribution date", "scheduled date"]),
    ("title", "Comms Title", 34,
     ["title", "subject", "name", "comms title", "message", "headline",
      "communication", "description", "task activity description",
      "activity description", "task description"]),
    ("purpose", "Purpose", 46,
     ["purpose", "objective", "why", "intent", "goal", "key message",
      "purpose / key message", "rationale"]),
    ("audience", "Audience", 30,
     ["audience", "target audience", "recipients", "who", "stakeholder",
      "stakeholders", "to", "audience group", "activity type", "comms to",
      "recipient group"]),
    ("channel", "Channel", 22,
     ["channel", "medium", "method", "vehicle", "delivery", "delivery method",
      "channel / medium"]),
    ("sender", "Sender (From)", 24,
     ["sender", "from", "issued by", "signed by", "sent by", "sender (from)",
      "on behalf of", "voice"]),
    ("owner", "Owner (Drafts/Sends)", 24,
     ["owner", "author", "drafter", "responsible", "prepared by", "accountable",
      "owner (drafts/sends)", "raci - r", "pic", "pic1", "person in charge",
      "tech support performing the task activity", "performing the task"]),
    ("approver", "Approver", 24,
     ["approver", "approval", "sign off", "sign-off", "signoff", "reviewed by",
      "approved by", "endorser", "verifier", "sign off by verifier"]),
    ("dependencies", "Dependencies", 38,
     ["dependency", "dependencies", "depends on", "pre-requisite", "prerequisite",
      "prerequisites", "inputs", "blockers", "linked activity"]),
    ("content_link", "Comms Content Link", 26,
     ["content", "content link", "comms content", "comms content link", "link",
      "collateral", "artefact", "artifact", "material", "draft link",
      "document link", "attachment"]),
    ("status", "Status", 14,
     ["status", "state", "progress", "rag"]),
    ("notes", "Notes", 32,
     ["notes", "comments", "remarks", "additional info"]),
]

FIELD_KEYS = [f[0] for f in FIELDS]
HEADERS = {f[0]: f[1] for f in FIELDS}
WIDTHS = {f[0]: f[2] for f in FIELDS}
DATE_FIELDS = {"send_date"}

# Fields that must exist as a column in the output even if a template lacks them.
REQUIRED_COLUMNS = ["purpose", "audience", "channel", "sender", "owner",
                    "approver", "dependencies", "content_link"]

# --------------------------------------------------------------------------
# Template profiles
# --------------------------------------------------------------------------
# A profile adapts the canonical plan to a client's own house format: their
# category vocabulary, their comms numbering, and any column of theirs that
# can be derived rather than left blank.

def _minus_business_days(value, days):
    """Back up from a send date to a draft-by date, skipping weekends."""
    if not isinstance(value, (datetime, date)):
        value = coerce_date(value)
    if not isinstance(value, (datetime, date)):
        return ""
    current = value.date() if isinstance(value, datetime) else value
    remaining = days
    while remaining > 0:
        current = current - timedelta(days=1)
        if current.weekday() < 5:
            remaining -= 1
    return current


# Milestone -> the client's Activity Category vocabulary.
ENG_CATEGORY = [
    (re.compile(r"^T-|readiness|action|chaser|programme", re.I), "Reminder Comms"),
    (re.compile(r"go/?no.?go|ponr|cutover begins|outage|service restored", re.I),
     "Cutover Period Comms"),
    (re.compile(r"mid-?cutover|checkpoint", re.I), "Checkpoint Comms"),
    (re.compile(r"go-?live|retired|complete", re.I), "Go Live Comms"),
    (re.compile(r"^T\+", re.I), "Post Go Live Comms"),
]

# One audience fragment -> the client's Activity Type vocabulary. Applied per
# fragment, not to the whole string: their plan carries one audience per row.
ENG_ACTIVITY_TYPE = [
    (re.compile(r"external|customer|vendor|partner|regulator|3rd party|third party|3p", re.I),
     "Comms to 3P"),
    (re.compile(r"exec|steer|board|c-suite|leadership|risk and compliance", re.I),
     "Comms to SC"),
    (re.compile(r"cutover team|programme|program team|bridge|service desk|duty|it ops", re.I),
     "Comms to {programme}"),
]


def eng_activity_type(fragment, programme):
    for pattern, label in ENG_ACTIVITY_TYPE:
        if pattern.search(fragment):
            return label.replace("{programme}", programme)
    return "Comms to BU"


def eng_split_audiences(text, programme):
    """Group a multi-audience string into the client's one-row-per-audience shape.

    Returns [(activity_type, original audience fragments)] preserving order, so
    'All Finance users; Service Desk; executive stakeholders' becomes three rows
    the way their own plan splits a milestone across BU / programme / Steerco.
    """
    fragments = [f.strip() for f in str(text or "").split(";") if f.strip()]
    if not fragments:
        return [("Comms to BU", "")]
    grouped = {}
    order = []
    for fragment in fragments:
        bucket = eng_activity_type(fragment, programme)
        if bucket not in grouped:
            grouped[bucket] = []
            order.append(bucket)
        grouped[bucket].append(fragment)
    return [(bucket, "; ".join(grouped[bucket])) for bucket in order]


# Their Status dropdown vocabulary, which differs from the canonical one —
# note "Disseminated" is their term for a comms that has gone out.
ENG_STATUS = {
    "not started": "Not Started", "drafting": "In Progress",
    "in review": "In Progress", "approved": "In Progress",
    "scheduled": "In Progress", "sent": "Disseminated",
    "cancelled": "Cancelled",
}


def _suffix(index):
    """a, b, ... z, aa, ab — so a comms with many audiences still numbers cleanly."""
    letters = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(97 + remainder) + letters
    return letters


def eng_transform(rows, spec, options):
    """Rewrite canonical rows into the ENG cutover task-list vocabulary.

    One comms addressed to several audiences becomes several rows, which is how
    this format is structured and what makes the #Na/#Nb suffixes meaningful.
    """
    programme = options.get("programme", "Programme Team")
    expanded = []
    for index, row in enumerate(rows, start=1):
        milestone = str(row.get("milestone", ""))
        category = next((label for pattern, label in ENG_CATEGORY
                         if pattern.search(milestone)), "Reminder Comms")
        for activity_type, detail in eng_split_audiences(row.get("audience"), programme):
            new_row = dict(row)
            new_row["_category"] = category
            new_row["_milestone"] = milestone
            new_row["audience"] = activity_type
            new_row["_audience_detail"] = detail
            # One number per comms, suffixed per audience — six checkpoint calls
            # are six comms, not one comms with six recipients.
            new_row["_seq"] = index
            expanded.append(new_row)

    # Where one milestone carries several rows, suffix them a/b/c.
    counts = {}
    for row in expanded:
        counts[row["_seq"]] = counts.get(row["_seq"], 0) + 1
    seen = {}
    for row in expanded:
        seq = row["_seq"]
        if counts[seq] > 1:
            seen[seq] = seen.get(seq, 0)
            row["id"] = "Comms #%d%s" % (seq, _suffix(seen[seq]))
            seen[seq] += 1
        else:
            row["id"] = "Comms #%d" % seq
        # Their taxonomy column replaces the canonical milestone on the way out.
        row["milestone"] = row["_category"]
        row["status"] = ENG_STATUS.get(str(row.get("status", "")).strip().lower(),
                                       row.get("status", ""))

    if len(expanded) != len(rows):
        print(f"  Audience split : {len(rows)} comms -> {len(expanded)} rows "
              f"(one per audience group, matching the #Na/#Nb convention)")
    rows[:] = expanded
    return rows


PROFILES = {
    "eng-cutover": {
        "description": "Cutover activity task list (Activity Category / Activity "
                       "Type / Comms # vocabulary, PIC and Verifier columns).",
        "signature": ["activity category", "task activity description"],
        # Pinned explicitly rather than left to synonym matching: this format has
        # two plausible id columns (Task # and Comms #) and a helper column whose
        # text trips loose matches.
        "columns": {
            "id": "Comms",
            "milestone": "Activity Category",
            "audience": "Activity Type",
            "title": "Task / Activity Description",
            "dependencies": "Dependency",
            "status": "Status",
            "owner": "Tech Support Performing the Task/Activity",
            "approver": "Sign-Off by Verifier (State the name of the Verifier)",
            "send_date": "Start Date",
            "notes": "Reference / Comments",
        },
        "transform": eng_transform,
        # Columns of theirs we can fill without being asked.
        "derived": {
            "Draft Created On": lambda row: _minus_business_days(row.get("send_date"), 3),
        },
        # Their Activity Type is a four-way bucket, so the specific group the
        # comms actually goes to would otherwise be lost.
        "extra_columns": [
            ("Audience (detail)", 30, lambda row: row.get("_audience_detail", "")),
        ],
        "vocabulary_check": ["milestone", "audience", "status"],
        # Client programme names stay out of this file — pass --programme, or let
        # the vocabulary snap pick up the template's own wording.
        "options": {"programme": "Programme Team"},
    },
}

CHANNEL_OPTIONS = [
    "Email", "Email (all-user)", "Teams/Slack post", "Teams/Slack channel",
    "Intranet article", "Manager cascade", "Town hall / all-hands",
    "Service desk banner", "In-app / system banner", "SMS / push",
    "Customer notice", "Vendor notice", "Poster / digital signage",
]

STATUS_OPTIONS = [
    "Not started", "Drafting", "In review", "Approved", "Scheduled", "Sent",
    "Cancelled",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
LABEL_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=10)
LINK_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def normalize(text):
    """Lowercase, strip punctuation/whitespace for fuzzy header matching."""
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def coerce_date(value):
    """Return a date object for ISO-ish strings, else the original value."""
    if isinstance(value, (datetime, date)):
        return value
    if not isinstance(value, str):
        return value
    text = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return value  # free text like "Cutover day, 06:00" is legitimate


def load_spec(path):
    with open(path, encoding="utf-8") as handle:
        spec = json.load(handle)
    if not isinstance(spec, dict):
        sys.exit("Spec must be a JSON object.")
    comms = spec.get("comms")
    if not isinstance(comms, list) or not comms:
        sys.exit("Spec must contain a non-empty 'comms' array.")
    unknown = {k for row in comms if isinstance(row, dict) for k in row} - set(FIELD_KEYS)
    if unknown:
        print(f"  ! Ignoring unrecognised comms keys: {', '.join(sorted(unknown))}",
              file=sys.stderr)
    return spec


def cell_value(field, row):
    value = row.get(field, "")
    if value is None:
        return ""
    if field in DATE_FIELDS:
        return coerce_date(value)
    return value


# --------------------------------------------------------------------------
# Mode 1: build from scratch
# --------------------------------------------------------------------------
def build_from_scratch(spec, out_path):
    workbook = Workbook()
    plan = workbook.active
    plan.title = "Comms Plan"

    for index, key in enumerate(FIELD_KEYS, start=1):
        cell = plan.cell(row=1, column=index, value=HEADERS[key])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        plan.column_dimensions[get_column_letter(index)].width = WIDTHS[key]
    plan.row_dimensions[1].height = 30

    for offset, row in enumerate(spec["comms"]):
        excel_row = offset + 2
        for index, key in enumerate(FIELD_KEYS, start=1):
            cell = plan.cell(row=excel_row, column=index, value=cell_value(key, row))
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in DATE_FIELDS and isinstance(cell.value, (datetime, date)):
                cell.number_format = "dd-mmm-yyyy"
            if key == "content_link":
                # Deliberately blank on generation - future linkage to content.
                cell.fill = LINK_FILL

    last_row = len(spec["comms"]) + 1
    plan.freeze_panes = "D2"
    plan.auto_filter.ref = f"A1:{get_column_letter(len(FIELD_KEYS))}{last_row}"

    _build_plan_info(workbook, spec)
    reference = _build_reference(workbook)
    _add_validations(plan, reference, last_row)

    workbook.save(out_path)
    return last_row - 1


def _build_plan_info(workbook, spec):
    sheet = workbook.create_sheet("Plan Info")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 78

    sheet["A1"] = spec.get("project") or "Cutover Communications Plan"
    sheet["A1"].font = TITLE_FONT

    rows = [
        ("Cutover type", spec.get("cutover_type", "")),
        ("Complexity tier", spec.get("complexity_tier", "")),
        ("Go-live date", spec.get("go_live_date", "")),
        ("Cutover window", spec.get("cutover_window", "")),
        ("Number of comms", len(spec["comms"])),
        ("Cadence rule applied", spec.get("cadence_rule", "")),
        ("Complexity modifiers applied", spec.get("modifiers", "")),
        ("Prepared by", spec.get("prepared_by", "")),
        ("Version", spec.get("version", "0.1 Draft")),
        ("Last updated", spec.get("last_updated", date.today().isoformat())),
        ("Notes", spec.get("notes", "")),
    ]
    for offset, (label, value) in enumerate(rows, start=3):
        if isinstance(value, list):
            value = "; ".join(str(item) for item in value)
        sheet.cell(row=offset, column=1, value=label).font = LABEL_FONT
        cell = sheet.cell(row=offset, column=2, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_reference(workbook):
    sheet = workbook.create_sheet("Reference")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 62
    sheet.column_dimensions["D"].width = 26
    sheet.column_dimensions["F"].width = 20

    sheet["A1"] = "Column dictionary"
    sheet["A1"].font = TITLE_FONT
    definitions = {
        "id": "Unique reference for the comms line item.",
        "milestone": "Cutover milestone the comms is anchored to (T-14, T-1, Cutover begins, Go-live, T+7).",
        "send_date": "Actual calendar date/time the comms goes out.",
        "title": "Working title or subject line.",
        "purpose": "What this comms is for - the decision, action, or awareness it drives.",
        "audience": "Who receives it. One audience group per line where the message differs.",
        "channel": "How it is delivered.",
        "sender": "Whose name it goes out under - the voice with the authority the message needs.",
        "owner": "Who drafts, schedules and sends it. Accountable for delivery.",
        "approver": "Who signs off before it goes out. Should not be the same person as the owner.",
        "dependencies": "What must be true or done first (go/no-go outcome, runbook step, distribution list, translations).",
        "content_link": "BLANK BY DESIGN - link to the drafted content once it exists.",
        "status": "Drafting/approval progress.",
        "notes": "Anything else.",
    }
    for offset, key in enumerate(FIELD_KEYS, start=3):
        sheet.cell(row=offset, column=1, value=HEADERS[key]).font = LABEL_FONT
        cell = sheet.cell(row=offset, column=2, value=definitions[key])
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet["D1"] = "Channel list"
    sheet["D1"].font = TITLE_FONT
    for offset, option in enumerate(CHANNEL_OPTIONS, start=3):
        sheet.cell(row=offset, column=4, value=option).font = BODY_FONT

    sheet["F1"] = "Status list"
    sheet["F1"].font = TITLE_FONT
    for offset, option in enumerate(STATUS_OPTIONS, start=3):
        sheet.cell(row=offset, column=6, value=option).font = BODY_FONT

    return sheet


def _add_validations(plan, reference, last_row):
    """Attach dropdowns to Channel and Status, sourced from the Reference sheet."""
    validation_rows = max(last_row, 200)
    specs = [
        ("channel", "D", len(CHANNEL_OPTIONS)),
        ("status", "F", len(STATUS_OPTIONS)),
    ]
    for key, ref_col, count in specs:
        column = get_column_letter(FIELD_KEYS.index(key) + 1)
        formula = f"='{reference.title}'!${ref_col}$3:${ref_col}${count + 2}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Pick a value from the Reference sheet list."
        validation.errorTitle = "Value not in list"
        plan.add_data_validation(validation)
        validation.add(f"{column}2:{column}{validation_rows}")


# --------------------------------------------------------------------------
# Mode 2: populate an existing template
# --------------------------------------------------------------------------
def find_header_row(sheet, forced=None, scan_rows=20):
    """Return (header_row_index, {field_key: column_index})."""
    best = (0, None, {})
    candidates = [forced] if forced else range(1, min(scan_rows, sheet.max_row or 1) + 1)
    for row_index in candidates:
        mapping = {}
        for column_index in range(1, (sheet.max_column or 1) + 1):
            header = normalize(sheet.cell(row=row_index, column=column_index).value)
            if not header:
                continue
            key = match_header(header)
            if key and key not in mapping:
                mapping[key] = column_index
        if len(mapping) > best[0]:
            best = (len(mapping), row_index, mapping)
    if forced:
        return forced, best[2]
    if best[0] < 3:
        return None, {}
    return best[1], best[2]


#: Substring matching on very short synonyms produces false positives — "to"
#: matches "val_[To Be Removed Later]", "no" matches "Notification". They stay
#: usable as exact matches only.
MIN_SUBSTRING_SYNONYM = 4


def match_header(normalized_header):
    """Map a normalized template header to a canonical field key."""
    for key, _, _, synonyms in FIELDS:
        if normalized_header in (normalize(s) for s in synonyms):
            return key
    # Fall back to substring containment, longest synonym first for precision.
    scored = []
    for key, _, _, synonyms in FIELDS:
        for synonym in synonyms:
            token = normalize(synonym)
            if len(token) >= MIN_SUBSTRING_SYNONYM and token in normalized_header:
                scored.append((len(token), key))
    if scored:
        scored.sort(reverse=True)
        return scored[0][1]
    return None


def detect_profile(sheet, header_row):
    """Return the profile key whose signature headers all appear in the template."""
    headers = {normalize(sheet.cell(row=header_row, column=c).value)
               for c in range(1, (sheet.max_column or 1) + 1)}
    for key, profile in PROFILES.items():
        if all(any(normalize(sig) == h or normalize(sig) in h for h in headers)
               for sig in profile["signature"]):
            return key
    return None


def count_populated(sheet, header_row, mapping):
    """Rows below the header that already carry data.

    Scans every column, not just the mapped ones: separator and end-of-plan
    marker rows often sit outside the mapped columns (and inside merged ranges),
    and appending on top of one both loses their content and hits a read-only
    merged cell.
    """
    populated = 0
    for row_index in range(header_row + 1, (sheet.max_row or header_row) + 1):
        if any(sheet.cell(row=row_index, column=c).value not in (None, "")
               for c in range(1, (sheet.max_column or 1) + 1)):
            populated = row_index - header_row
    return populated


def validation_list(sheet, column_index):
    """Values allowed by a dropdown on this column, if there is one.

    Resolves both literal lists and references to a range elsewhere in the
    workbook, so a template that documents its vocabulary on a separate sheet
    is as readable as one with the values inline.
    """
    letter = get_column_letter(column_index)
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list" or not validation.formula1:
            continue
        if not any(r.min_col <= column_index <= r.max_col for r in validation.sqref.ranges):
            continue
        formula = validation.formula1.strip().lstrip("=")
        if formula.startswith('"') and formula.endswith('"'):
            return [part.strip() for part in formula[1:-1].split(",") if part.strip()]
        match = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$", formula)
        if match:
            name, col_from, row_from, _, row_to = match.groups()
            if name not in sheet.parent.sheetnames:
                continue
            source = sheet.parent[name]
            values = []
            for row_index in range(int(row_from), int(row_to) + 1):
                value = source[f"{col_from}{row_index}"].value
                if value not in (None, ""):
                    values.append(str(value))
            return values
    return []


def snap_value(written, existing):
    """Match a generated value onto the template's own wording.

    Catches trailing-space variants ('Comms to BU ') and values a client has
    extended ('Comms to Programme' -> 'Comms to Programme (FIN, PROC, ...)'),
    either of which would otherwise read as a separate value in their filters.
    """
    target = written.strip().lower()
    for candidate in existing:
        if candidate.strip().lower() == target:
            return candidate
    if len(target) >= 6:
        prefixed = [c for c in existing if c.strip().lower().startswith(target)]
        if len(prefixed) == 1:
            return prefixed[0]
    return None


def merged_rows(sheet):
    """Row indices covered by any merged range, which cannot be written to."""
    covered = set()
    for merged in sheet.merged_cells.ranges:
        covered.update(range(merged.min_row, merged.max_row + 1))
    return covered


def strip_comments(workbook):
    """Remove cell comments anywhere in the workbook.

    Comments outlive the rows they discuss, are invisible in a quick scroll, and
    routinely name people and reference prior projects.
    """
    removed = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    text = re.sub(r"\s+", " ", str(cell.comment.text or "")).strip()
                    removed.append(f"{sheet.title}!{cell.coordinate}: {text[:70]}")
                    cell.comment = None
    return removed


def replace_text(workbook, replacements):
    """Substitute text across cell values, dropdown lists and sheet names.

    Project identifiers hide in all three, so a scrub that only walks cells
    leaves the vocabulary and the tab names still naming the old project.
    """
    counts = {old: 0 for old, _ in replacements}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for old, new in replacements:
                    if old in cell.value:
                        cell.value = cell.value.replace(old, new)
                        counts[old] += 1
        for validation in sheet.data_validations.dataValidation:
            formula = validation.formula1 or ""
            if not formula.strip().startswith('"'):
                continue
            for old, new in replacements:
                if old in formula:
                    validation.formula1 = formula = formula.replace(old, new)
                    counts[old] += 1
    for sheet in workbook.worksheets:
        for old, new in replacements:
            if old in sheet.title:
                sheet.title = sheet.title.replace(old, new)
                counts[old] += 1
    return counts


def parse_replacements(raw):
    """Parse OLD=>NEW pairs. '=>' rather than '=' so values may contain '='."""
    pairs = []
    for entry in raw or []:
        if "=>" not in entry:
            sys.exit(f"  ! Could not parse --replace-text {entry!r}. Expected OLD=>NEW.")
        old, new = entry.split("=>", 1)
        if not old:
            sys.exit("  ! --replace-text needs a non-empty search string.")
        pairs.append((old, new))
    return pairs


def apply_edits(workbook, default_sheet, drop_sheets, set_cells):
    """Drop sheets and overwrite individual cells during a template clean.

    The row clear only touches the data region, so anything project-specific in
    the title block, in group headers above the data, or on a reference sheet
    survives it. These stay as options rather than built-in rules because the
    wording is the client's, not ours.
    """
    for name in drop_sheets or []:
        matches = [s for s in workbook.sheetnames if s.strip().lower() == name.strip().lower()]
        if not matches:
            sys.exit(f"  ! No sheet named '{name}'. Sheets: {workbook.sheetnames}")
        if len(workbook.sheetnames) - len(matches) < 1:
            sys.exit("  ! Refusing to drop every sheet in the workbook.")
        for match in matches:
            del workbook[match]
            print(f"  Dropped sheet  : {match!r}")

    for assignment in set_cells or []:
        match = re.match(r"^(?:(.+)!)?([A-Za-z]{1,3}\d+)=(.*)$", assignment, re.S)
        if not match:
            sys.exit(f"  ! Could not parse --set-cell {assignment!r}. "
                     f"Expected [Sheet!]CELL=value, e.g. \"Plan!B1=[Project] Cutover\".")
        name, coordinate, value = match.groups()
        if name and name not in workbook.sheetnames:
            sys.exit(f"  ! No sheet named '{name}'. Sheets: {workbook.sheetnames}")
        target = workbook[name] if name else default_sheet
        cell = target[coordinate.upper()]
        if isinstance(cell, MergedCell):
            sys.exit(f"  ! {coordinate.upper()} on '{target.title}' is a merged cell. "
                     f"Set the top-left cell of the merged range instead.")
        previous = cell.value
        cell.value = value
        print(f"  Set            : {target.title}!{coordinate.upper()} = {value!r} "
              f"(was {str(previous)[:44]!r})")


def clean_template(template_path, out_path, sheet_name=None, header_row=None,
                   keep_through=None, drop_sheets=None, set_cells=None,
                   comments=False, replacements=None):
    """Strip a completed plan back to a blank template.

    Clients hand over last cutover's finished plan far more often than a blank
    form. This clears the data rows while keeping everything that makes the file
    a template — header block, column widths, styles, dropdowns, merged title
    ranges — and widens any list validation that only covered the old data, so
    the dropdown still works on rows the previous plan never reached.
    """
    workbook = load_workbook(template_path)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            sys.exit(f"Sheet '{sheet_name}' not found. Sheets: {workbook.sheetnames}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[0]

    header_row, mapping = find_header_row(sheet, forced=header_row)
    if not header_row:
        sys.exit("Could not identify a header row. Use --sheet and --header-row.")

    first = (keep_through or header_row) + 1
    last = count_populated(sheet, header_row, mapping) + header_row

    # The rows about to be cleared are the only record of the client's own
    # category vocabulary. Harvest it before it goes, and hand it back as
    # dropdowns so the blank template still documents its own taxonomy.
    profile = PROFILES.get(detect_profile(sheet, header_row) or "")
    harvested = {}
    if profile:
        for key in profile.get("vocabulary_check", []):
            column_index = mapping.get(key)
            if not column_index or validation_list(sheet, column_index):
                continue  # already has a dropdown — theirs wins
            values = []
            for row_index in range(header_row + 1, last + 1):
                value = sheet.cell(row=row_index, column=column_index).value
                if value not in (None, "") and str(value) not in values:
                    values.append(str(value))
            if values and len(values) <= 25:
                harvested[column_index] = (
                    str(sheet.cell(row=header_row, column=column_index).value or key).strip(),
                    values)
    print(f"  Sheet          : {sheet.title} (header row {header_row})")
    if last < first:
        print("  Nothing to clear — no populated rows below the header.")
        workbook.save(out_path)
        return 0

    # Merged ranges inside the data region are read-only and would survive as
    # stray blocks, so drop them; ranges in the header block stay.
    unmerged = []
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= first:
            unmerged.append(str(merged))
            sheet.unmerge_cells(str(merged))

    cleared, links, comment_count = 0, 0, 0
    for row_index in range(first, last + 1):
        for column_index in range(1, (sheet.max_column or 1) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value is not None:
                cell.value = None
            # A hyperlink outlives its display text and resurfaces as the cell's
            # value — and these routinely point at internal document stores with
            # a sharing token in the query string.
            if cell.hyperlink is not None:
                cell.hyperlink = None
                links += 1
            if cell.comment is not None:
                cell.comment = None
                comment_count += 1
        # Heights were sized to the old content; empty rows shouldn't keep them.
        if row_index in sheet.row_dimensions:
            sheet.row_dimensions[row_index].height = None
        cleared += 1
    sheet._hyperlinks = [h for h in sheet._hyperlinks
                         if not (first <= CellRange(h.ref).min_row <= last)]

    # A validation that only spanned the old rows leaves later rows unprotected.
    # Ranges are merged rather than appended: overlapping sqref entries make
    # Excel report the workbook as corrupt.
    widened = []
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        columns = {r.min_col for r in validation.sqref.ranges} | \
                  {r.max_col for r in validation.sqref.ranges}
        if len(columns) != 1:
            continue
        column = get_column_letter(columns.pop())
        intervals = sorted([(r.min_row, r.max_row) for r in validation.sqref.ranges]
                           + [(first, first + 199)])
        merged_intervals = []
        for start, end in intervals:
            if merged_intervals and start <= merged_intervals[-1][1] + 1:
                merged_intervals[-1][1] = max(merged_intervals[-1][1], end)
            else:
                merged_intervals.append([start, end])
        validation.sqref = MultiCellRange(
            [CellRange(f"{column}{s}:{column}{e}") for s, e in merged_intervals])
        widened.append(str(validation.sqref))

    # Values here routinely contain commas ("Comms to X (FIN, PROC, ...)"), which
    # a literal list formula would split, so the vocabulary goes on its own sheet
    # and the validations point at ranges.
    if harvested:
        name = "Vocabulary"
        if name in workbook.sheetnames:
            del workbook[name]
        vocabulary = workbook.create_sheet(name)
        vocabulary["A1"] = "Allowed values, harvested from the sample rows this template was built from."
        vocabulary["A1"].font = Font(bold=True, size=10)
        for offset, (column_index, (label, values)) in enumerate(sorted(harvested.items())):
            letter = get_column_letter(offset + 1)
            vocabulary.column_dimensions[letter].width = max(18, min(52, len(label) + 12))
            vocabulary.cell(row=3, column=offset + 1, value=label).font = Font(bold=True, size=10)
            for index, value in enumerate(values):
                vocabulary.cell(row=4 + index, column=offset + 1, value=value)
            validation = DataValidation(
                type="list", allow_blank=True,
                formula1=f"='{name}'!${letter}$4:${letter}${3 + len(values)}")
            sheet.add_data_validation(validation)
            source = get_column_letter(column_index)
            validation.add(f"{source}{first}:{source}{first + 199}")

    # Links to other workbooks survive the data that referenced them. They leak
    # internal document-store paths (often a named person's drive), and make
    # Excel prompt about external data every time the template is opened. Only
    # dropped when nothing could still reference them.
    has_formula = any(isinstance(c.value, str) and c.value.startswith("=")
                      for worksheet in workbook.worksheets
                      for row in worksheet.iter_rows() for c in row)
    dropped_links = 0
    if workbook._external_links:
        if has_formula:
            print(f"  ! {len(workbook._external_links)} external workbook link(s) kept — "
                  f"this workbook still contains formulas that may reference them.",
                  file=sys.stderr)
        else:
            dropped_links = len(workbook._external_links)
            workbook._external_links = []

    print(f"  Cleared        : rows {first}-{last} ({cleared} rows, all columns)")
    if links or comment_count:
        print(f"  Removed        : {links} hyperlink(s), {comment_count} comment(s) "
              f"attached to the cleared rows")
    if dropped_links:
        print(f"  Removed        : {dropped_links} orphaned external workbook link(s) "
              f"(no formulas reference them)")
    if harvested:
        summary = "; ".join(f"{label} ({len(values)})"
                            for label, values in sorted(harvested.values()))
        print(f"  Vocabulary kept: {summary} -> 'Vocabulary' sheet, wired up as dropdowns")
    if unmerged:
        print(f"  Unmerged       : {', '.join(unmerged)}")
    if widened:
        print(f"  Dropdown widened to cover the data region: {', '.join(widened)}")
    apply_edits(workbook, sheet, drop_sheets, set_cells)

    if comments:
        for entry in strip_comments(workbook):
            print(f"  Comment removed: {entry}")
    else:
        lingering = [f"{s.title}!{c.coordinate}"
                     for s in workbook.worksheets for r in s.iter_rows()
                     for c in r if c.comment is not None]
        if lingering:
            print(f"  ! {len(lingering)} comment(s) remain ({', '.join(lingering[:4])}) — "
                  f"they can name people and prior projects. Use --strip-comments.",
                  file=sys.stderr)

    if replacements:
        for old, count in replace_text(workbook, replacements).items():
            if count:
                print(f"  Replaced       : {old!r} in {count} place(s)")
            else:
                print(f"  ! --replace-text {old!r} matched nothing.", file=sys.stderr)

    remaining = ", ".join(repr(n) for n in workbook.sheetnames)
    print(f"  Kept           : rows 1-{header_row} (title block and headers), "
          f"column widths, cell styles, conditional formatting")
    print(f"  Sheets         : {remaining}")

    # Authorship is the client's call, so report rather than strip it.
    author = getattr(workbook.properties, "creator", None)
    modified_by = getattr(workbook.properties, "lastModifiedBy", None)
    named = ", ".join(n for n in {author, modified_by} if n and n != "openpyxl")
    if named:
        print(f"  Review         : document properties still name {named} — "
              f"clear them in File > Info if the template is going outside the team.")
    workbook.save(out_path)
    return cleared


def build_from_template(spec, template_path, out_path, sheet_name=None, header_row=None,
                        profile_key=None, mode="guard", options=None):
    workbook = load_workbook(template_path)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            sys.exit(f"Sheet '{sheet_name}' not found. Sheets: {workbook.sheetnames}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[0]

    header_row, mapping = find_header_row(sheet, forced=header_row)
    if not header_row or not mapping:
        sys.exit(
            "Could not identify a header row in the template. Re-run with "
            "--sheet and --header-row to point at it explicitly."
        )

    if profile_key is None:
        profile_key = detect_profile(sheet, header_row)
    profile = PROFILES.get(profile_key) if profile_key else None

    print(f"  Template sheet : {sheet.title} (header row {header_row})")
    if profile:
        print(f"  Profile        : {profile_key} — {profile['description']}")
        # Pinned columns win over anything synonym matching guessed.
        headers = {normalize(sheet.cell(row=header_row, column=c).value): c
                   for c in range(1, (sheet.max_column or 1) + 1)}
        for key, header_text in profile.get("columns", {}).items():
            column_index = headers.get(normalize(header_text))
            if column_index is None:
                print(f"  ! Profile expects a '{header_text}' column and this sheet "
                      f"has none — falling back to auto-matching for {key}.",
                      file=sys.stderr)
                continue
            previous = mapping.get(key)
            mapping[key] = column_index
            if previous and previous != column_index:
                print(f"  Pinned         : {key} -> {get_column_letter(column_index)} "
                      f"'{header_text}' (auto-match had {get_column_letter(previous)})")
        # Drop any auto-match that collided onto a pinned column.
        pinned = set(profile.get("columns", {}))
        for key in [k for k, c in mapping.items()
                    if k not in pinned and c in {mapping[p] for p in pinned if p in mapping}]:
            del mapping[key]
    matched = ", ".join(sorted(mapping))
    print(f"  Matched columns: {matched}")

    # Never write over a populated plan without being told to.
    populated = count_populated(sheet, header_row, mapping)
    start_row = header_row + 1
    if populated:
        if mode == "guard":
            sys.exit(
                f"  ! This template already holds {populated} populated row(s) below the "
                f"header.\n    Writing here would overwrite them. Choose explicitly:\n"
                f"      --append          keep them and add the new plan beneath\n"
                f"      --replace-rows    clear them and write the new plan in their place"
            )
        if mode == "append":
            start_row = header_row + 1 + populated
            print(f"  Existing rows  : {populated} kept, appending from row {start_row}")
        elif mode == "replace":
            protected = merged_rows(sheet)
            cleared, skipped = 0, []
            for row_index in range(header_row + 1, header_row + 1 + populated):
                if row_index in protected:
                    # Separator and marker rows live in merged ranges and are
                    # read-only; leaving them intact is also the right outcome.
                    skipped.append(row_index)
                    continue
                for column_index in mapping.values():
                    sheet.cell(row=row_index, column=column_index).value = None
                cleared += 1
            print(f"  Existing rows  : {cleared} cleared and replaced")
            if skipped:
                print(f"  Merged rows    : {', '.join(str(r) for r in skipped)} left "
                      f"intact (merged ranges cannot be cleared)")

    if profile and profile.get("transform"):
        merged_options = dict(profile.get("options", {}))
        merged_options.update(options or {})
        profile["transform"](spec["comms"], spec, merged_options)

    # Any required column the template lacks gets appended so no data is lost.
    next_column = (sheet.max_column or len(mapping)) + 1
    for key in REQUIRED_COLUMNS:
        if key in mapping:
            continue
        if not any(str(row.get(key, "")).strip() for row in spec["comms"]) \
                and key != "content_link":
            continue
        cell = sheet.cell(row=header_row, column=next_column, value=HEADERS[key])
        template_header = sheet.cell(row=header_row, column=1)
        if template_header.has_style:
            cell._style = copy(template_header._style)
        sheet.column_dimensions[get_column_letter(next_column)].width = WIDTHS[key]
        mapping[key] = next_column
        print(f"  + Added missing column '{HEADERS[key]}' at "
              f"{get_column_letter(next_column)}{header_row}")
        next_column += 1

    dropped = [HEADERS[k] for k in FIELD_KEYS
               if k not in mapping and any(str(r.get(k, "")).strip() for r in spec["comms"])]
    if dropped:
        print(f"  ! No template column for: {', '.join(dropped)} - this data is not written.",
              file=sys.stderr)

    # Profile-specific columns the template has no equivalent for.
    derived = {}
    if profile:
        for header_text, width, compute in profile.get("extra_columns", []):
            cell = sheet.cell(row=header_row, column=next_column, value=header_text)
            template_header = sheet.cell(row=header_row, column=1)
            if template_header.has_style:
                cell._style = copy(template_header._style)
            sheet.column_dimensions[get_column_letter(next_column)].width = width
            derived[next_column] = compute
            print(f"  + Added column   '{header_text}' at "
                  f"{get_column_letter(next_column)}{header_row}")
            next_column += 1

    # Columns of theirs the profile can fill without being asked (e.g. a
    # draft-by date derived from the send date and the approval lead time).
    if profile:
        for header_text, compute in profile.get("derived", {}).items():
            target = normalize(header_text)
            for column_index in range(1, (sheet.max_column or 1) + 1):
                if normalize(sheet.cell(row=header_row, column=column_index).value) == target:
                    derived[column_index] = compute
                    print(f"  Derived column : '{header_text}' at "
                          f"{get_column_letter(column_index)}")
                    break

    # Snap generated vocabulary onto whatever the template already uses, then
    # flag whatever genuinely has no precedent — so a new category value is a
    # decision rather than a surprise, and a near-miss doesn't become one.
    if profile:
        for key in profile.get("vocabulary_check", []):
            column_index = mapping.get(key)
            if not column_index:
                continue
            existing = []
            for row_index in range(header_row + 1, start_row):
                value = sheet.cell(row=row_index, column=column_index).value
                if value not in (None, "") and str(value) not in existing:
                    existing.append(str(value))
            allowed = validation_list(sheet, column_index)

            snapped = {}
            for row in spec["comms"]:
                written = str(row.get(key, "")).strip()
                if not written:
                    continue
                # A cleaned template has no rows to learn from, but may still
                # carry its vocabulary as a dropdown.
                match = snap_value(written, existing + [a for a in allowed if a not in existing])
                if match is not None and match != row.get(key):
                    snapped[written] = match
                    row[key] = match
            for source, target in sorted(snapped.items()):
                print(f"  Vocabulary     : {key} '{source}' -> '{target}' (template's own wording)")

            introduced = sorted({str(row.get(key, "")).strip() for row in spec["comms"]}
                                - {e.strip() for e in existing}
                                - {a.strip() for a in allowed} - {""})
            # A cleaned template has no rows but may carry a vocabulary, which
            # is just as good a precedent to measure a new value against.
            if introduced and (existing or allowed):
                print(f"  ! New {key} value(s) with no precedent in this template: "
                      f"{', '.join(introduced)}", file=sys.stderr)

    blocked = merged_rows(sheet)
    while start_row in blocked:
        print(f"  Row {start_row} sits inside a merged range — skipping it.")
        start_row += 1

    style_source = header_row + 1
    for offset, row in enumerate(spec["comms"]):
        excel_row = start_row + offset
        if excel_row in blocked:
            sys.exit(f"  ! Row {excel_row} is inside a merged range and cannot be "
                     f"written. Re-run with --append to write below the merged block.")
        for column_index, compute in derived.items():
            cell = sheet.cell(row=excel_row, column=column_index, value=compute(row))
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "dd-mmm-yyyy"
        for key, column_index in mapping.items():
            cell = sheet.cell(row=excel_row, column=column_index,
                              value=cell_value(key, row))
            source = sheet.cell(row=style_source, column=column_index)
            if excel_row != style_source and source.has_style:
                cell._style = copy(source._style)
            if key in DATE_FIELDS and isinstance(cell.value, (datetime, date)) \
                    and not cell.number_format.startswith(("d", "m", "y")):
                cell.number_format = "dd-mmm-yyyy"

    workbook.save(out_path)
    return len(spec["comms"])


# --------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--spec", help="Path to the JSON plan spec.")
    parser.add_argument("--out", help="Path to write the .xlsx to.")
    parser.add_argument("--template", help="Existing .xlsx template to populate.")
    parser.add_argument("--sheet", help="Sheet name within the template.")
    parser.add_argument("--header-row", type=int,
                        help="1-based header row in the template (skips auto-detection).")
    parser.add_argument("--profile", choices=sorted(PROFILES),
                        help="Client format profile. Auto-detected when omitted.")
    parser.add_argument("--no-profile", action="store_true",
                        help="Suppress profile auto-detection and map columns literally.")
    parser.add_argument("--programme", help="Programme name used in the profile's "
                                            "audience vocabulary (e.g. 'Comms to <Programme>').")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--append", action="store_true",
                       help="Keep the template's existing rows and add the plan beneath.")
    group.add_argument("--replace-rows", action="store_true",
                       help="Clear the template's existing rows and write the plan in place.")
    parser.add_argument("--clean-template", action="store_true",
                        help="Strip a completed plan back to a blank template "
                             "(no --spec needed) and write it to --out.")
    parser.add_argument("--keep-through", type=int,
                        help="With --clean-template, keep rows up to and including "
                             "this one (e.g. to preserve a sample first row).")
    parser.add_argument("--drop-sheet", action="append", metavar="NAME",
                        help="With --clean-template, delete this sheet. Repeatable.")
    parser.add_argument("--set-cell", action="append", metavar="[SHEET!]CELL=VALUE",
                        help="With --clean-template, overwrite a cell the row clear "
                             "does not reach — a title block or a group header. "
                             "Repeatable.")
    parser.add_argument("--strip-comments", action="store_true",
                        help="With --clean-template, remove every cell comment. "
                             "Comments are invisible when scrolling and routinely "
                             "name people and prior projects.")
    parser.add_argument("--replace-text", action="append", metavar="OLD=>NEW",
                        help="With --clean-template, substitute text across cell "
                             "values, dropdown lists and sheet names. Repeatable.")
    parser.add_argument("--list-fields", action="store_true",
                        help="Print the spec keys and exit.")
    parser.add_argument("--list-profiles", action="store_true",
                        help="Print the available template profiles and exit.")
    args = parser.parse_args()

    if args.list_fields:
        for key, header, _, _ in FIELDS:
            print(f"{key:<14} -> {header}")
        return
    if args.list_profiles:
        for key, profile in sorted(PROFILES.items()):
            print(f"{key}\n    {profile['description']}\n"
                  f"    detected by: {', '.join(profile['signature'])}")
        return

    if args.clean_template:
        if not args.template or not args.out:
            parser.error("--clean-template needs --template and --out")
        clean_template(args.template, args.out, args.sheet, args.header_row,
                       args.keep_through, args.drop_sheet, args.set_cell,
                       args.strip_comments, parse_replacements(args.replace_text))
        print(f"Wrote blank template to {args.out}")
        return

    if not args.spec or not args.out:
        parser.error("--spec and --out are required (or use --list-fields)")

    spec = load_spec(args.spec)
    if args.template:
        mode = "append" if args.append else "replace" if args.replace_rows else "guard"
        options = {"programme": args.programme} if args.programme else {}
        count = build_from_template(spec, args.template, args.out,
                                    args.sheet, args.header_row,
                                    # False suppresses detection; None triggers it.
                                    profile_key=False if args.no_profile else args.profile,
                                    mode=mode, options=options)
    else:
        count = build_from_scratch(spec, args.out)
    print(f"Wrote {count} comms line item(s) to {args.out}")


if __name__ == "__main__":
    main()
