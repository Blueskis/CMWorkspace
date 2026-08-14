#!/usr/bin/env python3
"""
Normalise a folder of mixed client source material into readable text + a source manifest.

Usage:
    python3 ingest_sources.py <source_dir> -o ingested/
    python3 ingest_sources.py <source_dir> -o ingested/ --transcribe    # audio too

Handles, with the standard library alone:
    .docx  .pptx           Word / PowerPoint text, headings, tables, speaker notes
    .xlsx .csv .tsv        sheet dumps (openpyxl for .xlsx)
    .vtt .srt              meeting transcripts — speaker turns with timestamps
    .txt .md .rtf          passthrough
    .bpmn .xml             Signavio/BPMN — pools, lanes, tasks, flows
    .json                  passthrough (pretty-printed)

Flags for the Read tool rather than extracting:
    .pdf .png .jpg .jpeg .gif .webp
    Claude reads these natively, including diagrams, scans and page layout, which a
    text extractor would throw away. The manifest lists them as read_natively.

Audio (.m4a .mp3 .wav .mp4 .aac .ogg .flac .wma) needs --transcribe and faster-whisper.
Without it they are listed as needs_transcript with instructions. See
../reference/source-ingestion.md for the audio decision — including why sending client
interview recordings to a cloud ASR service is a data-protection call, not a technical one.

Output:
    <out_dir>/<REF>__<slug>.md   one readable file per source, with a provenance header
    <out_dir>/sources.json       manifest — paste into meta.source_documents
    <out_dir>/INGEST_REPORT.md   what was read, what was skipped, what needs attention
"""

import argparse
import csv
import io
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import datetime

# --------------------------------------------------------------------------------------
# File classification. `ref_prefix` seeds the manifest; the analyst confirms the type —
# a .docx can be interview notes or a functional spec and only a human can tell.
# --------------------------------------------------------------------------------------

KINDS = {
    ".vtt": ("transcript", "INT", "Interview Transcript"),
    ".srt": ("transcript", "INT", "Interview Transcript"),
    ".docx": ("document", "DOC", "Document"),  # analyst reclassifies: notes? spec? org design?
    ".pptx": ("slides", "PPT", "Presentation"),
    ".xlsx": ("workbook", "DAT", "Spreadsheet"),
    ".xlsm": ("workbook", "DAT", "Spreadsheet"),
    ".csv": ("table", "DAT", "Spreadsheet"),
    ".tsv": ("table", "DAT", "Spreadsheet"),
    ".txt": ("text", "DOC", "Document"),
    ".md": ("text", "DOC", "Document"),
    ".rtf": ("text", "DOC", "Document"),
    ".json": ("text", "DOC", "Document"),
    ".bpmn": ("bpmn", "SIG", "Process Design"),
    ".xml": ("bpmn", "SIG", "Process Design"),
    ".pdf": ("read_natively", "DOC", "Document"),
    ".png": ("read_natively", "IMG", "Other"),
    ".jpg": ("read_natively", "IMG", "Other"),
    ".jpeg": ("read_natively", "IMG", "Other"),
    ".gif": ("read_natively", "IMG", "Other"),
    ".webp": ("read_natively", "IMG", "Other"),
    ".m4a": ("audio", "INT", "Interview Recording"),
    ".mp3": ("audio", "INT", "Interview Recording"),
    ".wav": ("audio", "INT", "Interview Recording"),
    ".mp4": ("audio", "INT", "Interview Recording"),
    ".aac": ("audio", "INT", "Interview Recording"),
    ".ogg": ("audio", "INT", "Interview Recording"),
    ".flac": ("audio", "INT", "Interview Recording"),
    ".wma": ("audio", "INT", "Interview Recording"),
}

SKIP_NAMES = {".ds_store", "thumbs.db", "desktop.ini"}
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def slug(name, limit=50):
    s = re.sub(r"[^\w\s-]", "", os.path.splitext(name)[0]).strip()
    s = re.sub(r"[\s_]+", "-", s).strip("-")
    return (s[:limit] or "source").lower()


# --------------------------------------------------------------------------------------
# Extractors — stdlib only, so this runs anywhere without a setup ritual
# --------------------------------------------------------------------------------------

def extract_docx(path):
    """Word text with headings and tables, in document order."""
    out, warnings = [], []
    with zipfile.ZipFile(path) as z:
        if "word/document.xml" not in z.namelist():
            return "", ["Not a readable .docx (no word/document.xml)."]
        root = ET.fromstring(z.read("word/document.xml"))
    body = root.find(f"{W}body")
    if body is None:
        return "", ["No document body found."]

    def para_text(p):
        return "".join(t.text or "" for t in p.iter(f"{W}t"))

    def para_style(p):
        pr = p.find(f"{W}pPr")
        if pr is None:
            return None
        st = pr.find(f"{W}pStyle")
        return st.get(f"{W}val") if st is not None else None

    for el in body:
        tag = el.tag
        if tag == f"{W}p":
            txt = para_text(el).strip()
            if not txt:
                continue
            style = (para_style(el) or "").lower()
            m = re.match(r"heading(\d)", style)
            if m:
                out.append(f"\n{'#' * min(int(m.group(1)) + 1, 6)} {txt}\n")
            elif style.startswith("list"):
                out.append(f"- {txt}")
            else:
                out.append(txt)
        elif tag == f"{W}tbl":
            rows = []
            for tr in el.findall(f"{W}tr"):
                cells = ["  ".join(para_text(p).strip() for p in tc.findall(f"{W}p")).strip()
                         for tc in tr.findall(f"{W}tc")]
                rows.append(cells)
            if rows:
                out.append("")
                width = max(len(r) for r in rows)
                for i, r in enumerate(rows):
                    r = r + [""] * (width - len(r))
                    out.append("| " + " | ".join(c.replace("\n", " ") for c in r) + " |")
                    if i == 0:
                        out.append("|" + "---|" * width)
                out.append("")
    if not out:
        warnings.append("Document parsed but no text found — check it isn't a scan.")
    return "\n".join(out), warnings


def extract_pptx(path):
    """Slide text plus speaker notes. Notes routinely carry the reasoning the slide omits."""
    out, warnings = [], []
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        slides = sorted((n for n in names if re.match(r"ppt/slides/slide\d+\.xml$", n)),
                        key=lambda n: int(re.search(r"(\d+)", n).group(1)))
        if not slides:
            return "", ["No slides found."]
        for n in slides:
            idx = int(re.search(r"(\d+)", n).group(1))
            root = ET.fromstring(z.read(n))
            texts = [t.text.strip() for t in root.iter(f"{A}t") if (t.text or "").strip()]
            out.append(f"\n## Slide {idx}\n")
            out.extend(texts or ["_(no text on slide)_"])
            note = f"ppt/notesSlides/notesSlide{idx}.xml"
            if note in names:
                nroot = ET.fromstring(z.read(note))
                ntexts = [t.text.strip() for t in nroot.iter(f"{A}t") if (t.text or "").strip()]
                ntexts = [t for t in ntexts if t != str(idx)]  # drop the slide-number placeholder
                if ntexts:
                    out.append("\n**Speaker notes:** " + " ".join(ntexts))
    return "\n".join(out), warnings


def extract_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "", ["openpyxl not installed — cannot read .xlsx (pip install openpyxl)."]
    out, warnings = [], []
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        out.append(f"\n## Sheet: {ws.title}\n")
        n = 0
        for row in ws.iter_rows(values_only=True):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            out.append(" | ".join("" if c is None else str(c).replace("\n", " ") for c in row))
            n += 1
            if n >= 400:
                out.append(f"\n_… truncated at 400 populated rows._")
                warnings.append(f"Sheet '{ws.title}' truncated at 400 rows.")
                break
    wb.close()
    return "\n".join(out), warnings


def extract_delimited(path):
    delim = "\t" if path.lower().endswith(".tsv") else ","
    text, _ = read_text_file(path)
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    out, warnings = [], []
    for i, r in enumerate(rows[:400]):
        out.append(" | ".join(r))
    if len(rows) > 400:
        out.append("\n_… truncated at 400 rows._")
        warnings.append("Truncated at 400 rows.")
    return "\n".join(out), warnings


def read_text_file(path):
    raw = open(path, "rb").read()
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return raw.decode(enc), []
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), ["Encoding not detected; decoded with replacement."]


def extract_rtf(path):
    """Crude but adequate RTF de-controlling — enough to read notes pasted out of Word."""
    text, warnings = read_text_file(path)
    text = re.sub(r"\\'([0-9a-fA-F]{2})", lambda m: chr(int(m.group(1), 16)), text)
    text = re.sub(r"\\par[d]?\b", "\n", text)
    text = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", text)
    text = text.replace("{", "").replace("}", "")
    return re.sub(r"\n{3,}", "\n\n", text).strip(), warnings + [
        "RTF stripped with a simple parser — check for lost formatting."]


# ---- transcripts -----------------------------------------------------------------------

TS = r"(\d{1,2}:\d{2}:\d{2})[.,]\d{1,3}"


def extract_transcript(path):
    """
    Normalise .vtt/.srt into speaker turns with a start timestamp per turn.

    Handles the two conventions in the wild:
      Teams  — <v Jane Doe>text</v>
      Zoom   — Jane Doe: text
    Consecutive cues from the same speaker are merged, so a turn reads as a sentence
    rather than as caption fragments. Timestamps are kept because they are the citation:
    `INT-04 @00:23:15` lets a reviewer jump straight to the audio.
    """
    text, warnings = read_text_file(path)
    blocks = re.split(r"\n\s*\n", text.replace("\r\n", "\n").strip())
    turns = []
    for b in blocks:
        lines = [l for l in b.split("\n") if l.strip()]
        if not lines or lines[0].strip().upper().startswith("WEBVTT"):
            continue
        start = None
        body = []
        for line in lines:
            m = re.search(TS + r"\s*-->\s*" + TS, line)
            if m:
                start = m.group(1)
                continue
            if re.fullmatch(r"\d+", line.strip()) and start is None:
                continue  # SRT cue index
            body.append(line.strip())
        if not body:
            continue
        content = " ".join(body)
        speaker = None
        mv = re.match(r"<v\s+([^>]+?)>(.*)", content, re.S)
        if mv:
            speaker, content = mv.group(1).strip(), mv.group(2)
        else:
            ms = re.match(r"([A-Z][\w.'\- ]{1,40}):\s+(.*)", content, re.S)
            if ms:
                speaker, content = ms.group(1).strip(), ms.group(2)
        content = re.sub(r"</?v[^>]*>", "", content)
        content = re.sub(r"<[^>]+>", "", content).strip()
        if content:
            turns.append((start, speaker, content))

    if not turns:
        return "", ["No transcript cues recognised — check the file is .vtt or .srt."]

    merged = []
    for start, speaker, content in turns:
        if merged and merged[-1][1] == speaker and speaker is not None:
            merged[-1][2] += " " + content
        else:
            merged.append([start, speaker, content])

    speakers = sorted({s for _t, s, _c in merged if s})
    out = []
    if speakers:
        out.append(f"**Speakers detected:** {', '.join(speakers)}\n")
    else:
        warnings.append("No speaker labels found — attribution is unavailable, which lowers "
                        "the confidence of anything drawn from this transcript.")
    out.append(f"**Turns:** {len(merged)}   **Words:** ~{sum(len(c.split()) for _t,_s,c in merged):,}\n")
    out.append("---\n")
    for start, speaker, content in merged:
        stamp = f"[{start}] " if start else ""
        who = f"**{speaker}:** " if speaker else ""
        out.append(f"{stamp}{who}{content}\n")
    return "\n".join(out), warnings


# ---- BPMN / Signavio -------------------------------------------------------------------

def _local(tag):
    return tag.split("}")[-1]


def extract_bpmn(path):
    """
    Pull the structure a CIA actually needs out of a BPMN export: pools, lanes, the tasks
    in each lane, and the sequence flow. Lanes are the impacted roles — that mapping is
    the single most useful thing a process model gives a change impact assessment.
    """
    try:
        root = ET.parse(path).getroot()
    except ET.ParseError as e:
        return "", [f"Not parseable XML: {e}"]

    if not any(_local(el.tag) in ("definitions", "process", "collaboration") for el in root.iter()):
        return "", ["XML parsed but contains no BPMN elements — not a process model?"]

    names = {}
    for el in root.iter():
        eid = el.get("id")
        if eid:
            names[eid] = el.get("name") or ""

    TASKISH = {"task", "userTask", "serviceTask", "manualTask", "scriptTask", "sendTask",
               "receiveTask", "businessRuleTask", "subProcess", "callActivity"}
    EVENTISH = {"startEvent", "endEvent", "intermediateCatchEvent", "intermediateThrowEvent",
                "boundaryEvent"}
    GATEWAYISH = {"exclusiveGateway", "parallelGateway", "inclusiveGateway", "eventBasedGateway",
                  "complexGateway"}

    out, warnings = [], []

    participants = [el for el in root.iter() if _local(el.tag) == "participant"]
    if participants:
        out.append("## Pools\n")
        for p in participants:
            out.append(f"- **{p.get('name') or '(unnamed)'}**")
        out.append("")

    for proc in [el for el in root.iter() if _local(el.tag) == "process"]:
        pname = proc.get("name") or proc.get("id") or "(unnamed process)"
        out.append(f"\n## Process: {pname}\n")

        lane_of = {}
        lanes = [el for el in proc.iter() if _local(el.tag) == "lane"]
        for lane in lanes:
            lname = lane.get("name") or "(unnamed lane)"
            for ref in lane.iter():
                if _local(ref.tag) == "flowNodeRef" and (ref.text or "").strip():
                    lane_of[ref.text.strip()] = lname

        if lanes:
            out.append("### Lanes (→ impacted roles)\n")
            for lane in lanes:
                out.append(f"- **{lane.get('name') or '(unnamed lane)'}**")
            out.append("")
        else:
            warnings.append(f"Process '{pname}' has no lanes — impacted roles must come from elsewhere.")

        by_lane = defaultdict(list)
        for el in proc.iter():
            t = _local(el.tag)
            if t in TASKISH or t in EVENTISH or t in GATEWAYISH:
                by_lane[lane_of.get(el.get("id"), "(no lane)")].append(
                    (t, el.get("name") or "(unnamed)", el.get("id")))
        if by_lane:
            out.append("### Activities by lane\n")
            for lane in sorted(by_lane):
                out.append(f"\n**{lane}**\n")
                for t, name, eid in by_lane[lane]:
                    out.append(f"- `{t}` {name}")
            out.append("")

        flows = [el for el in proc.iter() if _local(el.tag) == "sequenceFlow"]
        if flows:
            out.append("### Sequence flow\n")
            for f in flows:
                src = names.get(f.get("sourceRef"), "") or f.get("sourceRef") or "?"
                tgt = names.get(f.get("targetRef"), "") or f.get("targetRef") or "?"
                cond = f" *[{f.get('name')}]*" if f.get("name") else ""
                out.append(f"- {src} → {tgt}{cond}")
            out.append("")

        docs = [(el.text or "").strip() for el in proc.iter()
                if _local(el.tag) == "documentation" and (el.text or "").strip()]
        if docs:
            out.append("### Documentation notes\n")
            out.extend(f"- {d}" for d in docs)

    if not out:
        return "", ["No pools, lanes or activities found."]
    return "\n".join(out), warnings


# ---- audio -----------------------------------------------------------------------------

def transcribe_audio(path, model_size="small", language=None):
    """
    Local transcription via faster-whisper. Local by design: interview recordings contain
    named employees discussing job changes, and shipping them to a third-party ASR service
    is a data-protection decision, not a convenience. See reference/source-ingestion.md.
    """
    try:
        from faster_whisper import WhisperModel
    except ImportError:
        return None, ["faster-whisper not installed. `pip install faster-whisper`, or supply "
                      "the meeting platform's own transcript instead (preferred — see "
                      "reference/source-ingestion.md)."]
    try:
        model = WhisperModel(model_size, device="cpu", compute_type="int8")
        segments, info = model.transcribe(path, language=language, vad_filter=True)
    except Exception as e:
        return None, [f"Transcription failed: {type(e).__name__}: {e}"]

    lines, count, words = [], 0, 0
    for seg in segments:
        h, rem = divmod(int(seg.start), 3600)
        m, s = divmod(rem, 60)
        txt = seg.text.strip()
        if txt:
            lines.append(f"[{h:02d}:{m:02d}:{s:02d}] {txt}")
            count += 1
            words += len(txt.split())

    header = [
        f"**Transcribed locally** with faster-whisper `{model_size}` — "
        f"detected language `{info.language}` (confidence {info.language_probability:.2f}), "
        f"duration {info.duration/60:.1f} min, {count} segments, ~{words:,} words.\n",
        "> **No speaker labels.** Machine transcription cannot tell who is speaking. Attribute "
        "turns yourself from the audio or the attendee list before relying on anything here — "
        "who said a thing determines whether it is testimony, a claim, or an opinion.\n",
        "> **Do not take figures from this transcript unquoted.** Headcounts, system names and "
        "acronyms are exactly what ASR gets wrong. Corroborate against a document or mark the "
        "row Low confidence.\n",
        "---\n",
    ]
    return "\n".join(header + lines), [
        "Audio transcribed by machine — no speaker attribution; verify names, acronyms and figures."]


# --------------------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------------------

EXTRACTORS = {
    "document": extract_docx,
    "slides": extract_pptx,
    "workbook": extract_xlsx,
    "table": extract_delimited,
    "bpmn": extract_bpmn,
    "transcript": extract_transcript,
}


def extract_plain(path):
    text, warnings = read_text_file(path)
    if path.lower().endswith(".json"):
        try:
            text = json.dumps(json.loads(text), indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            warnings.append("Not valid JSON — passed through as text.")
    return text, warnings


def ingest(source_dir, out_dir, do_transcribe, model_size, language):
    os.makedirs(out_dir, exist_ok=True)
    files = []
    for root, dirs, names in os.walk(source_dir):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for n in sorted(names):
            if n.lower() in SKIP_NAMES or n.startswith("~$") or n.startswith("."):
                continue
            files.append(os.path.join(root, n))
    files.sort()

    counters = defaultdict(int)
    manifest, report = [], []

    for path in files:
        ext = os.path.splitext(path)[1].lower()
        rel = os.path.relpath(path, source_dir)
        if ext not in KINDS:
            report.append({"file": rel, "status": "unsupported", "detail": f"extension {ext or '(none)'}"})
            continue

        kind, prefix, doc_type = KINDS[ext]
        counters[prefix] += 1
        ref = f"{prefix}-{counters[prefix]:02d}"
        warnings, text, status = [], None, "ok"

        if kind == "read_natively":
            status = "read_natively"
        elif kind == "audio":
            if not do_transcribe:
                status = "needs_transcript"
            else:
                text, warnings = transcribe_audio(path, model_size, language)
                status = "transcribed" if text else "transcription_failed"
        elif kind in ("text",):
            text, warnings = extract_plain(path)
        elif ext == ".rtf":
            text, warnings = extract_rtf(path)
        else:
            try:
                text, warnings = EXTRACTORS[kind](path)
            except Exception as e:
                text, warnings, status = None, [f"{type(e).__name__}: {e}"], "extraction_failed"

        entry = {
            "ref": ref,
            "type": doc_type,
            "title": os.path.splitext(os.path.basename(path))[0],
            "date": "",
            "author_or_participants": "",
            "_source_file": rel,
            "_status": status,
        }

        if text:
            words = len(text.split())
            if not text.strip():
                status = entry["_status"] = "empty"
                warnings.append("Extraction produced no text — likely a scan or an image-only file.")
            out_name = f"{ref}__{slug(os.path.basename(path))}.md"
            with open(os.path.join(out_dir, out_name), "w", encoding="utf-8") as fh:
                fh.write(f"# {ref} — {entry['title']}\n\n")
                fh.write(f"- **Source file:** `{rel}`\n")
                fh.write(f"- **Kind:** {kind}\n")
                fh.write(f"- **Extracted:** {datetime.now():%Y-%m-%d %H:%M}\n")
                if warnings:
                    fh.write("- **Warnings:**\n")
                    for w in warnings:
                        fh.write(f"  - {w}\n")
                fh.write("\n---\n\n")
                fh.write(text.rstrip() + "\n")
            entry["_extracted_to"] = out_name
            entry["_words"] = words

        entry["_warnings"] = warnings
        manifest.append(entry)
        report.append({"file": rel, "ref": ref, "status": entry["_status"],
                       "words": entry.get("_words"), "warnings": warnings})

    write_report(out_dir, source_dir, manifest, report, do_transcribe)
    clean = [{k: v for k, v in e.items() if not k.startswith("_")}
             for e in manifest if e["_status"] in ("ok", "transcribed", "read_natively")]
    with open(os.path.join(out_dir, "sources.json"), "w", encoding="utf-8") as fh:
        json.dump({"source_documents": clean}, fh, indent=2, ensure_ascii=False)
    return manifest, report


def write_report(out_dir, source_dir, manifest, report, do_transcribe):
    by_status = defaultdict(list)
    for r in report:
        by_status[r["status"]].append(r)

    L = [f"# Ingestion Report\n",
         f"Source: `{source_dir}`  \nRun: {datetime.now():%Y-%m-%d %H:%M}\n",
         f"**{len(report)} files seen.** "
         + ", ".join(f"{len(v)} {k}" for k, v in sorted(by_status.items())) + "\n"]

    ok = by_status["ok"] + by_status["transcribed"]
    if ok:
        L.append("\n## Extracted — read these\n")
        L.append("| Ref | File | Words | Notes |")
        L.append("|---|---|---:|---|")
        for r in ok:
            L.append(f"| `{r['ref']}` | `{r['file']}` | {r.get('words') or 0:,} | "
                     f"{'; '.join(r['warnings']) or '—'} |")

    if by_status["read_natively"]:
        L.append("\n## Read these with the Read tool, not a script\n")
        L.append("PDFs and images keep information a text extractor discards — page layout, "
                 "process diagrams, org charts, scanned annotations. Claude reads them natively.\n")
        L.append("| Ref | File |")
        L.append("|---|---|")
        for r in by_status["read_natively"]:
            L.append(f"| `{r['ref']}` | `{r['file']}` |")

    if by_status["needs_transcript"]:
        L.append("\n## Audio — needs a transcript before it can be used\n")
        L.append("Claude cannot listen to audio. These need transcribing first. In order of "
                 "preference:\n")
        L.append("1. **The meeting platform's own transcript.** Teams, Zoom, Meet and most "
                 "note-takers already produce one, with speaker labels and correctly spelled "
                 "names because the platform knows the attendee list. Export it as `.vtt` or "
                 "`.docx`, drop it in the source folder, and re-run. This is the best option "
                 "and usually costs nothing.\n")
        L.append("2. **Local machine transcription:** `pip install faster-whisper`, then re-run "
                 "with `--transcribe`. Stays on the machine, but produces no speaker labels.\n")
        L.append("3. A cloud ASR service — a data-protection decision to take deliberately, not "
                 "a default. See `reference/source-ingestion.md`.\n")
        L.append("| Ref | File |")
        L.append("|---|---|")
        for r in by_status["needs_transcript"]:
            L.append(f"| `{r['ref']}` | `{r['file']}` |")

    trouble = (by_status["extraction_failed"] + by_status["transcription_failed"]
               + by_status["empty"] + by_status["unsupported"])
    if trouble:
        L.append("\n## Needs attention\n")
        L.append("| File | Status | Detail |")
        L.append("|---|---|---|")
        for r in trouble:
            detail = r.get("detail") or "; ".join(r.get("warnings") or []) or "—"
            L.append(f"| `{r['file']}` | {r['status']} | {detail} |")
        L.append("\nAn unreadable source is a gap in the assessment, not a technicality — "
                 "say so at handover rather than letting it disappear.")

    L.append("\n## Next\n")
    L.append("1. Open `sources.json` and **correct the `type` and `title` of every entry.** The "
             "script guesses from the file extension; only a person can tell interview notes "
             "from a functional specification. Add dates and participants while you are there.\n")
    L.append("2. Read every extracted file in full — see `reference/interview-evidence.md` for "
             "reading transcripts, which is a different job from reading notes.\n")
    L.append("3. Build the register per `reference/extraction-guide.md`, citing these refs in "
             "`source_ref`.\n")

    with open(os.path.join(out_dir, "INGEST_REPORT.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(L) + "\n")


def main():
    ap = argparse.ArgumentParser(
        description="Normalise mixed client source material into text + a source manifest.")
    ap.add_argument("source_dir")
    ap.add_argument("-o", "--out", default="ingested")
    ap.add_argument("--transcribe", action="store_true",
                    help="Transcribe audio locally with faster-whisper (slow; prefer the "
                         "meeting platform's own transcript where one exists)")
    ap.add_argument("--model", default="small",
                    help="faster-whisper model: tiny/base/small/medium/large-v3 (default small)")
    ap.add_argument("--language", default=None, help="ISO code, e.g. en. Default: auto-detect")
    args = ap.parse_args()

    if not os.path.isdir(args.source_dir):
        print(f"✗ Not a directory: {args.source_dir}")
        return 1

    manifest, report = ingest(args.source_dir, args.out, args.transcribe, args.model, args.language)

    counts = defaultdict(int)
    for r in report:
        counts[r["status"]] += 1
    print(f"\n✓ {len(report)} files seen → {args.out}/")
    for status in sorted(counts):
        print(f"    {status:22} {counts[status]}")
    words = sum(r.get("words") or 0 for r in report)
    if words:
        print(f"\n  ~{words:,} words extracted")
    if counts.get("needs_transcript"):
        print(f"\n  ⚠  {counts['needs_transcript']} audio file(s) need a transcript — "
              f"see {args.out}/INGEST_REPORT.md")
    if counts.get("read_natively"):
        print(f"  ℹ  {counts['read_natively']} PDF/image(s) to open with the Read tool")
    print(f"\n  Next: correct types and titles in {args.out}/sources.json, then read "
          f"{args.out}/INGEST_REPORT.md\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
