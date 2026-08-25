#!/usr/bin/env python3
"""
Extract impact rows from a filled-in CIA Template workbook.

Usage:
    python3 import_cia_excel.py "Client CIA.xlsx" -o extracted.json
    python3 import_cia_excel.py "Client CIA.xlsx" --sheet "CIA Template"

Reads whatever's populated in the client CIA template's own sheet — the default
columns, or the --extended governance columns if present — and emits a JSON list of
raw row dicts using the same keys as cia_input.json's impacts[] (see
../reference/input-schema.md).

This is extraction only, same division of labour as ingest_sources.py for every
other input format: it does not score anything the sheet left blank and does not
push to Airtable. Claude reads the output, applies Step 4 (score) / Step 5 (derive
response) to whatever the sheet left incomplete — exactly as it would for a
free-text brief or a submitted intake form — then writes cia_input.json and/or
pushes to Airtable.

Column layout is imported directly from generate_cia.py (TEMPLATE_COLUMNS /
EXTENDED_COLUMNS) so the two scripts can never drift apart on what a column means.
Matching is by position, not by exact header text — generate_cia.py's own column
map treats header strings as "for reference only" (it writes by index), and two of
the template's real header cells carry a full instructional sentence after the
short label (As-is / To-be). A `startswith`-after-whitespace-normalising check
catches a genuinely reordered or deleted column while tolerating that.
"""
import argparse
import json
import os
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from generate_cia import TEMPLATE_COLUMNS, EXTENDED_COLUMNS, TEMPLATE_SHEET  # noqa: E402

CONTENT_KEYS = {
    "l1", "l2", "l3", "l4", "current_roles", "as_is", "to_be",
    "people_impact", "process_impact", "tech_impact",
}


def norm(s):
    return " ".join(str(s or "").split()).strip().lower()


def find_header_row(ws):
    for r in range(1, 6):
        if norm(ws.cell(row=r, column=1).value) == "l1":
            return r
    raise ValueError(
        "Could not find the header row (looking for a cell reading exactly 'L1' "
        "in column A, within rows 1-5). Is this the CIA Template sheet?"
    )


def detect_columns(ws, header_row):
    """Match the header row against TEMPLATE_COLUMNS positionally, then check
    whether EXTENDED_COLUMNS follows. Raises ValueError with a readable diff if
    the base columns have been reordered, renamed or deleted."""
    mismatches = []
    for i, (key, expected, _w, _kind) in enumerate(TEMPLATE_COLUMNS):
        actual = ws.cell(row=header_row, column=i + 1).value
        if not norm(actual).startswith(norm(expected)):
            mismatches.append(f"  column {i + 1} ({key}): expected to start with "
                               f"'{expected}', found {actual!r}")
    if mismatches:
        raise ValueError(
            "This sheet's columns don't match the CIA Template layout:\n"
            + "\n".join(mismatches)
            + "\n\nUse the vendored templates/CIA_Template.xlsx, or a file "
              "generate_cia.py produced, without reordering, renaming or "
              "deleting columns."
        )

    base_n = len(TEMPLATE_COLUMNS)
    ext_found = [
        ws.cell(row=header_row, column=base_n + i + 1).value
        for i in range(len(EXTENDED_COLUMNS))
    ]
    is_extended = bool(any(ext_found)) and all(
        norm(actual).startswith(norm(expected))
        for (_key, expected, _w, _kind), actual in zip(EXTENDED_COLUMNS, ext_found)
    )
    cols = list(TEMPLATE_COLUMNS) + (list(EXTENDED_COLUMNS) if is_extended else [])
    return cols, is_extended


def coerce(value, kind):
    if value is None:
        return None
    if kind in ("text", "wrap"):
        s = str(value).strip()
        return s or None
    if kind in ("int", "code", "score"):
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return int(value)
        s = str(value).strip()
        return int(s) if s.lstrip("-").isdigit() else None
    return None  # "formula" (Overall Impact) is recomputed downstream, never imported


def is_blank_row(raw_values, cols):
    for (key, *_r), v in zip(cols, raw_values):
        if key in CONTENT_KEYS and v not in (None, ""):
            return False
    return True


def extract(path, sheet_name=None, max_blank_run=5):
    wb = load_workbook(path, data_only=True)
    sheet_name = sheet_name or TEMPLATE_SHEET
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Sheets in this file: "
            f"{', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]
    header_row = find_header_row(ws)
    cols, is_extended = detect_columns(ws, header_row)

    rows = []
    blank_run = 0
    r = header_row + 1
    while blank_run < max_blank_run and r <= ws.max_row:
        raw_values = [ws.cell(row=r, column=i + 1).value for i in range(len(cols))]
        if is_blank_row(raw_values, cols):
            blank_run += 1
            r += 1
            continue
        blank_run = 0
        row_dict = {"_row": r}
        for (key, _header, _w, kind), v in zip(cols, raw_values):
            if kind == "formula":
                continue
            row_dict[key] = coerce(v, kind)
        rows.append(row_dict)
        r += 1

    return rows, is_extended, header_row


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("workbook")
    ap.add_argument("--sheet", default=None,
                     help=f"Sheet name (default: '{TEMPLATE_SHEET}')")
    ap.add_argument("-o", "--output", default=None,
                     help="Write JSON here instead of stdout")
    args = ap.parse_args()

    try:
        rows, is_extended, header_row = extract(args.workbook, args.sheet)
    except (ValueError, FileNotFoundError) as e:
        print(f"error: {e}", file=sys.stderr)
        sys.exit(1)

    complete = sum(
        1 for r in rows
        if all(r.get(k) is not None for k in ("score_people", "score_process", "score_technology"))
    )
    print(
        f"# {len(rows)} row(s) extracted from '{args.sheet or TEMPLATE_SHEET}' "
        f"(header row {header_row}, {'extended' if is_extended else 'default'} columns)",
        file=sys.stderr,
    )
    print(
        f"# {complete}/{len(rows)} row(s) already carry all three dimension scores; "
        f"the rest need Step 4 (score) before they can be pushed to Airtable.",
        file=sys.stderr,
    )

    out = json.dumps(rows, indent=2, ensure_ascii=False)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(out)
        print(f"# wrote {args.output}", file=sys.stderr)
    else:
        print(out)


if __name__ == "__main__":
    main()
