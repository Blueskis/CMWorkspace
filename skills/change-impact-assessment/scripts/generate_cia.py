#!/usr/bin/env python3
"""
Populate the client CIA template with a baseline change impact assessment.

Usage:
    python3 generate_cia.py cia_input.json -o "Change Impact Assessment.xlsx"
    python3 generate_cia.py cia_input.json --extended     # + governance columns
    python3 generate_cia.py cia_input.json --validate-only

Input contract: see ../reference/input-schema.md
Rating model:   see ../reference/rating-methodology.md  (client template's own 0-3 scale)

The workbook is built by loading templates/CIA_Template.xlsx and writing rows into
the 'CIA Template' sheet from row 3. The template's headers, theme colours, merges
and its 'Change Impact Ratings' rubric sheet are carried through untouched.
Supplementary sheets (heatmap, training, comms, traceability, assessment info) are
appended after the two template sheets.
"""

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(HERE, "..", "templates", "CIA_Template.xlsx")

TEMPLATE_SHEET = "CIA Template"
RUBRIC_SHEET = "Change Impact Ratings"
DATA_START = 3          # template headers occupy rows 1-2
BUFFER_ROWS = 50        # blank pre-formatted rows for adding impacts in-workbook

# --------------------------------------------------------------------------------------
# Rating model — the client template's own scale (see RUBRIC_SHEET in the workbook)
# --------------------------------------------------------------------------------------
# Three dimensions, each scored 0-3 against the anchors on the rubric sheet.
# Overall Impact = unweighted AVERAGE of the three, exactly as the template's header says.

DIMENSIONS = [
    ("score_people", "People"),
    ("score_process", "Process"),
    ("score_technology", "Technology"),
]

SCALE = {0: "No change", 1: "Low", 2: "Medium", 3: "High"}

# Band cut-offs for the 0-3 average. NOTE: the template defines the per-dimension scale
# but not the cut-offs for the overall average — these are this skill's assumption and
# are stated as such on the Assessment Info sheet. Change them here if the client has
# their own convention.
BANDS = [  # (lower bound inclusive, label)
    (2.5, "High"),
    (1.5, "Medium"),
    (0.5, "Low"),
    (0.0, "No / Minimal"),
]
RATINGS = ["No / Minimal", "Low", "Medium", "High"]

ENUMS = {
    "resistance_risk": ["Low", "Medium", "High"],
    "training_required": ["Yes", "No"],
    "comms_required": ["Yes", "No"],
    "training_type": [
        "Classroom ILT", "Virtual ILT", "e-Learning", "Job Aid", "In-App Guidance",
        "Floorwalking / Hypercare", "Webinar", "Not Required",
    ],
    "confidence": ["High", "Medium", "Low"],
    "validation_status": ["Draft", "In Review", "Validated", "Baselined"],
    "rating_override": RATINGS,
}

COMMS_WAVES = [
    ("Awareness", -12, -8, "Something is coming, and why"),
    ("Understanding", -8, -4, "What specifically changes for you"),
    ("Readiness", -4, -1, "What you must do, and when"),
    ("Go-Live", 0, 2, "Where to get help"),
    ("Reinforcement", 2, 12, "It's working, and it's permanent"),
]

# --------------------------------------------------------------------------------------
# Column map — mirrors the template's own header row exactly, in order
# --------------------------------------------------------------------------------------
# (json_key, template header (for reference only), column width, kind)
# kind: text | wrap | int | code | score | formula
TEMPLATE_COLUMNS = [
    ("l1", "L1", 20, "text"),
    ("l1_code", "L1 code", 8, "code"),
    ("l2", "L2", 24, "text"),
    ("l2_code", "L2 code", 8, "code"),
    ("l3", "L3", 28, "text"),
    ("l3_code", "L3 code", 8, "code"),
    ("l4", "L4", 32, "text"),
    ("l4_code", "L4 code", 8, "code"),
    ("current_roles", "Current Roles", 30, "wrap"),
    ("headcount_impacted", "Approx Headcount", 12, "int"),
    ("as_is", "As-is", 55, "wrap"),
    ("to_be", "To-be", 55, "wrap"),
    ("people_impact", "People", 44, "wrap"),
    ("score_people", "Degree of Impact", 10, "score"),
    ("process_impact", "Process", 44, "wrap"),
    ("score_process", "Degree of Impact", 10, "score"),
    ("tech_impact", "Tech", 44, "wrap"),
    ("score_technology", "Degree of Impact", 10, "score"),
    ("overall", "Overall Impact (Average)", 12, "formula"),
    ("training", "Training (Link to Training Plan)", 42, "wrap"),
    ("communications", "Communications (Link to Comms Plan)", 42, "wrap"),
    ("others", "Others (e.g., policies, engagements)", 44, "wrap"),
]

# Appended only with --extended. Keeps the client template exact by default.
EXTENDED_COLUMNS = [
    ("impact_id", "Impact ID", 11, "text"),
    ("stakeholder_group", "Stakeholder Group", 26, "text"),
    ("resistance_risk", "Anticipated Resistance", 13, "text"),
    ("change_champion", "Change Champion", 24, "wrap"),
    ("source_ref", "Source Document Ref", 18, "text"),
    ("confidence", "Confidence", 11, "text"),
    ("validation_status", "Validation Status", 14, "text"),
    ("notes", "Notes / Open Questions", 45, "wrap"),
]

REQUIRED_IMPACT_FIELDS = [
    "impact_id", "l1", "l2", "l3", "current_roles", "as_is", "to_be",
    "people_impact", "process_impact", "tech_impact",
    "score_people", "score_process", "score_technology",
    "stakeholder_group", "resistance_risk", "training_required", "comms_required",
    "source_ref", "confidence",
]
REQUIRED_META_FIELDS = [
    "program_name", "client", "solution_scope", "assessment_owner",
    "version", "assessment_date", "source_documents",
]
SCORE_FIELDS = [k for k, _ in DIMENSIONS]

# --------------------------------------------------------------------------------------
# Styling — matched to the template's theme palette
# --------------------------------------------------------------------------------------

TPL = "'CIA Template'"

# Rubric sheet's own colour language: Low = light green, Medium = amber, High = light orange
FILL_NONE = PatternFill("solid", fgColor="F2F2F2")
FILL_LOW = PatternFill("solid", fgColor="B8DCAB")
FILL_MED = PatternFill("solid", fgColor="FFDB69")
FILL_HIGH = PatternFill("solid", fgColor="F6C6AD")

RATING_FILLS = {
    "No / Minimal": FILL_NONE,
    "Low": FILL_LOW,
    "Medium": FILL_MED,
    "High": FILL_HIGH,
}
RATING_FONTS = {
    "No / Minimal": Font(name="Calibri", size=10, color="7F7F7F"),
    "Low": Font(name="Calibri", size=10, bold=True, color="2C5F2D"),
    "Medium": Font(name="Calibri", size=10, bold=True, color="7F6000"),
    "High": Font(name="Calibri", size=10, bold=True, color="9C4221"),
}
SCORE_FILLS = {0: FILL_NONE, 1: FILL_LOW, 2: FILL_MED, 3: FILL_HIGH}

HEAD_FILL = PatternFill("solid", fgColor=Color(theme=1, tint=0.25))
HEAD_FONT = Font(name="Calibri", size=11, color=Color(theme=0))
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=Color(theme=0))
BODY = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
SUBTLE = Font(name="Calibri", size=9, italic=True, color="666666")

THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STRIPE = PatternFill("solid", fgColor="F7F9FB")
BAND = PatternFill("solid", fgColor="E8EDF2")


def columns(extended):
    return TEMPLATE_COLUMNS + (EXTENDED_COLUMNS if extended else [])


def col_letter(key, extended=False):
    for i, (k, *_r) in enumerate(columns(extended), start=1):
        if k == key:
            return get_column_letter(i)
    raise KeyError(key)


# --------------------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------------------

def validate(data):
    errors, warnings = [], []

    meta = data.get("meta")
    if not isinstance(meta, dict):
        return ["`meta` object is missing."], []

    for f in REQUIRED_META_FIELDS:
        if not meta.get(f):
            errors.append(f"meta.{f} is required but missing or empty.")

    known_refs = set()
    for i, d in enumerate(meta.get("source_documents") or []):
        ref = (d or {}).get("ref")
        if not ref:
            errors.append(f"meta.source_documents[{i}] has no `ref`.")
        else:
            if ref in known_refs:
                errors.append(f"Duplicate source document ref '{ref}'.")
            known_refs.add(ref)

    impacts = data.get("impacts")
    if not isinstance(impacts, list) or not impacts:
        errors.append("`impacts` must be a non-empty array.")
        return errors, warnings

    seen_ids, cited_refs = set(), set()
    l1_ratings = defaultdict(list)

    for idx, imp in enumerate(impacts):
        tag = imp.get("impact_id") or f"impacts[{idx}]"

        for f in REQUIRED_IMPACT_FIELDS:
            v = imp.get(f)
            if v is None or (isinstance(v, str) and not v.strip()):
                errors.append(f"{tag}: required field `{f}` is missing or empty.")

        iid = imp.get("impact_id")
        if iid:
            if iid in seen_ids:
                errors.append(f"Duplicate impact_id '{iid}'.")
            seen_ids.add(iid)

        for f in SCORE_FIELDS:
            v = imp.get(f)
            if v is None:
                continue
            if isinstance(v, bool) or not isinstance(v, int):
                errors.append(f"{tag}: `{f}` must be an integer 0-3, got {v!r}.")
            elif not 0 <= v <= 3:
                errors.append(f"{tag}: `{f}` must be between 0 and 3 (template scale), got {v}.")

        for f, allowed in ENUMS.items():
            v = imp.get(f)
            if v and v not in allowed:
                errors.append(f"{tag}: `{f}` = {v!r} is not one of: {', '.join(allowed)}.")

        if imp.get("rating_override") and not (imp.get("rating_override_reason") or "").strip():
            errors.append(f"{tag}: `rating_override` set but `rating_override_reason` is empty.")

        for ref in split_refs(imp.get("source_ref")):
            cited_refs.add(ref)
            if known_refs and ref not in known_refs:
                errors.append(f"{tag}: source_ref '{ref}' is not declared in meta.source_documents.")

        # ---- warnings
        rating = computed_rating(imp)
        if imp.get("l1"):
            l1_ratings[imp["l1"]].append(rating)

        if imp.get("confidence") == "Low" and not (imp.get("notes") or "").strip():
            warnings.append(f"{tag}: Low confidence with no open question recorded in `notes`.")
        if imp.get("resistance_risk") == "High" and not (imp.get("mitigation_actions") or "").strip():
            warnings.append(f"{tag}: High anticipated resistance with no `mitigation_actions`.")
        if rating == "High" and not (imp.get("change_champion") or "").strip():
            warnings.append(f"{tag}: rated High with no `change_champion` named.")
        if rating == "High" and not (imp.get("comms_owner") or "").strip():
            warnings.append(f"{tag}: rated High with no `comms_owner` named.")
        if imp.get("training_required") == "Yes" and not imp.get("training_type"):
            warnings.append(f"{tag}: training required but no `training_type` set.")
        if imp.get("training_required") == "Yes" and not imp.get("training_duration_hrs"):
            warnings.append(f"{tag}: training required but no `training_duration_hrs` — effort cannot be rolled up.")
        if imp.get("comms_required") == "Yes" and not (imp.get("key_message") or "").strip():
            warnings.append(f"{tag}: comms required but no `key_message` drafted.")
        if rating == "High" and imp.get("training_required") == "No":
            warnings.append(f"{tag}: rated High but no training planned — check the rating or the response.")
        if not (imp.get("l4") or "").strip():
            warnings.append(f"{tag}: no L4 activity recorded — the template's taxonomy goes to four levels.")

    for l1, ratings in l1_ratings.items():
        if "High" not in ratings:
            warnings.append(f"L1 '{l1}' has no High-rated impacts — confirm this is real, not under-assessed.")

    for ref in sorted(known_refs - cited_refs):
        warnings.append(f"Source document '{ref}' is not cited by any impact — was it read?")

    return errors, warnings


def split_refs(value):
    """Split a source_ref into document refs, dropping any `@timestamp` locator.

    Transcript citations carry the moment as well as the document — `INT-04 @00:23:15` —
    so a reviewer can jump straight to the audio. Only the document part is validated.
    """
    if not value:
        return []
    out = []
    for part in str(value).replace(",", ";").split(";"):
        ref = part.split("@")[0].strip()
        if ref:
            out.append(ref)
    return out


def average_score(imp):
    try:
        return round(sum(imp[f] for f in SCORE_FIELDS) / len(SCORE_FIELDS), 2)
    except (KeyError, TypeError):
        return None


def computed_rating(imp):
    """Python mirror of the workbook's band logic, used for stats and warnings."""
    if imp.get("rating_override"):
        return imp["rating_override"]
    s = average_score(imp)
    if s is None:
        return None
    for lower, label in BANDS:
        if s >= lower:
            return label
    return "No / Minimal"


# --------------------------------------------------------------------------------------
# Composed template columns
# --------------------------------------------------------------------------------------

def compose_training(imp):
    if imp.get("training_required") != "Yes":
        return "Not required"
    bits = [imp.get("training_module_ref"), imp.get("training_type")]
    if imp.get("training_duration_hrs"):
        bits.append(f"{imp['training_duration_hrs']:g} hrs")
    aud = imp.get("training_audience_size") or imp.get("headcount_impacted")
    if aud:
        bits.append(f"{aud:,} learners")
    bits.append(imp.get("training_timing"))
    return " · ".join(b for b in bits if b)


def compose_comms(imp):
    if imp.get("comms_required") != "Yes":
        return "Not required"
    bits = []
    if imp.get("comms_timing"):
        bits.append(imp["comms_timing"])
    if imp.get("comms_channel"):
        bits.append(imp["comms_channel"])
    if imp.get("comms_owner"):
        bits.append(f"Owner: {imp['comms_owner']}")
    return " · ".join(bits)


def compose_others(imp):
    parts = []
    if (imp.get("other_impacts") or "").strip():
        parts.append(imp["other_impacts"].strip())
    if (imp.get("mitigation_actions") or "").strip():
        parts.append(f"Mitigation: {imp['mitigation_actions'].strip()}")
    if imp.get("rating_override"):
        parts.append(f"RATING OVERRIDE → {imp['rating_override']}: {imp.get('rating_override_reason','')}")
    return "\n\n".join(parts)


COMPOSED = {
    "training": compose_training,
    "communications": compose_comms,
    "others": compose_others,
}


# --------------------------------------------------------------------------------------
# Populate the template sheet
# --------------------------------------------------------------------------------------

def fill_template_sheet(wb, data, extended):
    ws = wb[TEMPLATE_SHEET]
    impacts = data["impacts"]
    cols = columns(extended)
    ncols = len(cols)
    last_data = DATA_START + len(impacts) - 1
    last_live = last_data + BUFFER_ROWS

    if extended:
        _add_extended_headers(ws, len(TEMPLATE_COLUMNS))

    for i, (key, _hdr, width, _kind) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ov = col_letter("overall", extended)
    sp, spr, st = (col_letter(f, extended) for f in SCORE_FIELDS)

    for r, imp in enumerate(impacts, start=DATA_START):
        for i, (key, _hdr, _w, kind) in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i)
            cell.border = CELL_BORDER
            cell.font = BODY
            if kind == "formula":
                continue
            value = COMPOSED[key](imp) if key in COMPOSED else imp.get(key)
            if key == "validation_status" and not value:
                value = "Draft"
            if kind == "code" and isinstance(value, str) and value.isdigit():
                value = int(value)
            cell.value = value
            if kind == "wrap":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif kind == "score":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, int):
                    cell.fill = SCORE_FILLS[value]
                    cell.font = BODY_BOLD
            elif kind == "int":
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.number_format = "#,##0"
            elif kind == "code":
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.number_format = "00"
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 108

    # -- overall average formula across data + buffer rows
    overrides = {DATA_START + i: imp.get("rating_override")
                 for i, imp in enumerate(impacts) if imp.get("rating_override")}
    for r in range(DATA_START, last_live + 1):
        c = ws.cell(row=r, column=cols.index(next(x for x in cols if x[0] == "overall")) + 1)
        c.value = (f'=IF(COUNT({sp}{r},{spr}{r},{st}{r})<3,"",'
                   f'ROUND(AVERAGE({sp}{r},{spr}{r},{st}{r}),2))')
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = BODY_BOLD
        for i in range(1, ncols + 1):
            cc = ws.cell(row=r, column=i)
            cc.border = CELL_BORDER
            if r > last_data:
                cc.font = BODY
                if cols[i - 1][3] == "wrap":
                    cc.alignment = Alignment(wrap_text=True, vertical="top")
                elif cols[i - 1][3] in ("score", "int", "code"):
                    cc.alignment = Alignment(horizontal="center", vertical="center")

    # zebra striping on populated rows, skipping the colour-coded score/overall columns
    coloured = {sp, spr, st, ov}
    for r in range(DATA_START, last_data + 1, 2):
        for i in range(1, ncols + 1):
            if get_column_letter(i) not in coloured:
                ws.cell(row=r, column=i).fill = STRIPE

    # -- conditional formatting: overall average, in the rubric sheet's colour language
    rng = f"{ov}{DATA_START}:{ov}{last_live}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["2.5"], fill=FILL_HIGH, font=RATING_FONTS["High"]))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["1.5", "2.49"], fill=FILL_MED, font=RATING_FONTS["Medium"]))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["0.5", "1.49"], fill=FILL_LOW, font=RATING_FONTS["Low"]))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0.5"], fill=FILL_NONE, font=RATING_FONTS["No / Minimal"]))

    # buffer rows: colour the degree cells as they're typed
    for letter in (sp, spr, st):
        brng = f"{letter}{last_data + 1}:{letter}{last_live}"
        for score, fill in SCORE_FILLS.items():
            ws.conditional_formatting.add(brng, CellIsRule(
                operator="equal", formula=[str(score)], fill=fill))

    # -- dropdowns for the 0-3 degree columns
    dv = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
    dv.errorTitle, dv.error = "Value not allowed", "Degree of Impact is 0-3 — see the Change Impact Ratings sheet."
    ws.add_data_validation(dv)
    for letter in (sp, spr, st):
        dv.add(f"{letter}{DATA_START}:{letter}{last_live}")

    if extended:
        for field in ("resistance_risk", "confidence", "validation_status"):
            d = DataValidation(type="list", formula1=f'"{",".join(ENUMS[field])}"', allow_blank=True)
            ws.add_data_validation(d)
            L = col_letter(field, extended)
            d.add(f"{L}{DATA_START}:{L}{last_live}")
        conf = col_letter("confidence", extended)
        ws.conditional_formatting.add(f"{conf}{DATA_START}:{conf}{last_live}", CellIsRule(
            operator="equal", formula=['"Low"'], fill=FILL_HIGH, font=RATING_FONTS["High"]))
        res = col_letter("resistance_risk", extended)
        for label, fill in (("High", FILL_HIGH), ("Medium", FILL_MED), ("Low", FILL_LOW)):
            ws.conditional_formatting.add(f"{res}{DATA_START}:{res}{last_live}", CellIsRule(
                operator="equal", formula=[f'"{label}"'], fill=fill))

    ws.freeze_panes = f"I{DATA_START}"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last_live}"
    ws.print_title_rows = "1:2"
    ws.sheet_view.zoomScale = 85

    # honour any override by writing the label into Others (already composed) — the
    # numeric average stays untouched so the template's own arithmetic remains auditable
    return last_data, last_live, overrides


def _add_extended_headers(ws, base_cols):
    first = base_cols + 1
    last = base_cols + len(EXTENDED_COLUMNS)
    ws.merge_cells(start_row=1, start_column=first, end_row=1, end_column=last)
    t = ws.cell(row=1, column=first, value="Governance & Traceability")
    t.fill, t.font = HEAD_FILL, TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for offset, (_k, header, _w, _kind) in enumerate(EXTENDED_COLUMNS):
        c = ws.cell(row=2, column=first + offset, value=header)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# --------------------------------------------------------------------------------------
# Supplementary sheets
# --------------------------------------------------------------------------------------

def style_title(ws, row, text, span, size=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=Color(theme=0))
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 26


def section_header(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True)
    c.fill = BAND
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def table_header(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = CELL_BORDER
    ws.row_dimensions[row].height = 30


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def rng(field, last_live, extended):
    L = col_letter(field, extended)
    return f"{TPL}!${L}${DATA_START}:${L}${last_live}"


def is_column(field, extended):
    return any(k == field for k, *_r in columns(extended))


def lookup(field, imp, id_cell, last_live, extended, row_fallback):
    """Live reference when the field is a workbook column; static value otherwise.

    Fields such as key_message and benefit_narrative live only in the JSON — they have
    no column in the client template, so the roll-up carries the value directly.
    When --extended is on, Impact ID exists and lookups are sort-safe INDEX/MATCH;
    otherwise they are direct row references into the register.
    """
    if field in COMPOSED:
        return COMPOSED[field](imp)
    if not is_column(field, extended):
        return imp.get(field, "")
    src = rng(field, last_live, extended)
    if extended:
        ids = rng("impact_id", last_live, extended)
        return f'=IFERROR(INDEX({src},MATCH({id_cell},{ids},0)),"")'
    return f"={TPL}!${col_letter(field, extended)}${row_fallback}"


BAND_CRITERIA = {
    "No / Minimal": ('"<0.5"', None),
    "Low": ('">=0.5"', '"<1.5"'),
    "Medium": ('">=1.5"', '"<2.5"'),
    "High": ('">=2.5"', None),
}


def band_countifs(key_range, key_cell, ov_range, label):
    lo, hi = BAND_CRITERIA[label]
    parts = [f"{key_range},{key_cell}", f"{ov_range},{lo}"]
    if hi:
        parts.append(f"{ov_range},{hi}")
    return "=COUNTIFS(" + ",".join(parts) + ")"


def build_heatmap(wb, data, last_live, extended):
    ws = wb.create_sheet("Impact Heatmap")
    impacts = data["impacts"]
    groups = sorted({i["stakeholder_group"] for i in impacts})
    l1s = sorted({i["l1"] for i in impacts})

    widths = [36, 13, 11, 12, 11, 11, 16, 18]
    set_widths(ws, widths)
    span = len(widths)

    ov = rng("overall", last_live, extended)
    grp = rng("stakeholder_group", last_live, extended) if extended else None
    l1r = rng("l1", last_live, extended)
    hc = rng("headcount_impacted", last_live, extended)

    style_title(ws, 1, "IMPACT HEATMAP — where the change actually lands", span)
    n = ws.cell(row=2, column=1, value=(
        "Live formulas over the CIA Template sheet. Re-score a Degree of Impact there and this "
        "view updates. Bands on the 0-3 average: High ≥ 2.5 · Medium 1.5–2.49 · Low 0.5–1.49 · "
        "No/Minimal < 0.5."))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    n.font, n.alignment = SUBTLE, Alignment(wrap_text=True, vertical="center", indent=1)

    row = 4

    def static_cells(field, key, label=None):
        """Counts computed in Python, for axes with no column in the client template."""
        sel = [i for i in impacts if i[field] == key]
        if label is None:
            return len(sel), sum(i.get("headcount_impacted") or 0 for i in sel), (
                round(sum(average_score(i) for i in sel) / len(sel), 2) if sel else None)
        return sum(1 for i in sel if computed_rating(i) == label)

    def matrix(title, keys, key_range, field=None):
        nonlocal row
        live = key_range is not None
        section_header(ws, row, title + ("" if live else "  (snapshot — see note below)"), span)
        row += 1
        table_header(ws, row, ["", "No / Minimal", "Low", "Medium", "High", "Total",
                               "Approx headcount", "Mean overall score"])
        row += 1
        first = row
        for key in keys:
            kc = ws.cell(row=row, column=1, value=key)
            kc.font, kc.border = BODY_BOLD, CELL_BORDER
            kc.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            for ci, label in enumerate(RATINGS, start=2):
                c = ws.cell(row=row, column=ci,
                            value=(band_countifs(key_range, f"$A{row}", ov, label) if live
                                   else static_cells(field, key, label)))
                c.number_format = "#,##0;;–"
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font, c.border = BODY, CELL_BORDER
            cnt, ppl, mean = (None, None, None) if live else static_cells(field, key)
            for ci, formula, fmt in (
                    (6, f"=SUM(B{row}:E{row})", "#,##0"),
                    (7, f"=SUMIFS({hc},{key_range},$A{row})" if live else ppl, "#,##0;;–"),
                    (8, (f'=IFERROR(ROUND(AVERAGEIFS({ov},{key_range},$A{row}),2),"–")' if live
                         else mean), "0.00")):
                c = ws.cell(row=row, column=ci, value=formula)
                c.number_format = fmt
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = BODY_BOLD if ci in (6, 8) else BODY
                c.border = CELL_BORDER
            ws.row_dimensions[row].height = 20
            row += 1
        tc = ws.cell(row=row, column=1, value="TOTAL")
        tc.font, tc.fill, tc.border = BODY_BOLD, BAND, CELL_BORDER
        tc.alignment = Alignment(indent=1)
        overall_mean = round(sum(average_score(i) for i in impacts) / len(impacts), 2)
        for ci in range(2, span + 1):
            L = get_column_letter(ci)
            if ci == 8:
                v = f'=IFERROR(ROUND(AVERAGE({ov}),2),"–")' if live else overall_mean
            else:
                v = f"=SUM({L}{first}:{L}{row - 1})"
            c = ws.cell(row=row, column=ci, value=v)
            c.font, c.fill, c.border = BODY_BOLD, BAND, CELL_BORDER
            c.alignment = Alignment(horizontal="center")
            c.number_format = "0.00" if ci == 8 else "#,##0"
        for ci, label in enumerate(RATINGS, start=2):
            L = get_column_letter(ci)
            ws.conditional_formatting.add(f"{L}{first}:{L}{row - 1}", CellIsRule(
                operator="greaterThan", formula=["0"],
                fill=RATING_FILLS[label], font=RATING_FONTS[label]))
        ws.conditional_formatting.add(f"H{first}:H{row - 1}", CellIsRule(
            operator="greaterThanOrEqual", formula=["2.5"], fill=FILL_HIGH))
        row += 2

    matrix("BY STAKEHOLDER GROUP", groups, grp, field="stakeholder_group")
    matrix("BY L1 PROCESS AREA", l1s, l1r, field="l1")

    if grp is None:
        note = ws.cell(row=row, column=1, value=(
            "Note: Stakeholder Group has no column in the client template, so the first table above is "
            "a snapshot computed at generation time rather than live formulas — re-generate after "
            "re-scoring, or run with --extended to add the column and make it live. The L1 table and "
            "every other sheet are live against the register."))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        note.font, note.alignment = SUBTLE, Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[row].height = 40

    ws.sheet_view.showGridLines = False


def build_rollup(wb, data, last_live, extended, *, sheet_name, title, blurb,
                 filter_field, fields, widths, summary=None):
    ws = wb.create_sheet(sheet_name)
    rows = [(DATA_START + i, imp) for i, imp in enumerate(data["impacts"])
            if imp.get(filter_field) == "Yes"]
    span = len(widths)
    set_widths(ws, widths)

    style_title(ws, 1, title, span)
    note = ws.cell(row=2, column=1, value=blurb + (
        "" if extended else "  Rows follow the register's row order — re-run the generator if you "
        "re-sort the CIA Template sheet (or use --extended for sort-safe lookups)."))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    note.font, note.alignment = SUBTLE, Alignment(wrap_text=True, vertical="center", indent=1)
    ws.row_dimensions[2].height = 26

    row = 4
    if summary:
        row = summary(ws, row, span) + 1

    section_header(ws, row, f"DETAIL — {len(rows)} impacts", span)
    row += 1
    table_header(ws, row, [h for _f, h in fields])
    row += 1
    first = row
    for src_row, imp in rows:
        idc = ws.cell(row=row, column=1, value=imp["impact_id"])
        idc.font = BODY_BOLD
        for ci, (field, _h) in enumerate(fields[1:], start=2):
            c = ws.cell(row=row, column=ci,
                        value=lookup(field, imp, f"$A{row}", last_live, extended, src_row))
            c.font = BODY
            if field == "overall":
                c.number_format = "0.00"
                c.font = BODY_BOLD
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "resistance_risk":
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        for i in range(1, span + 1):
            ws.cell(row=row, column=i).border = CELL_BORDER
        ws.row_dimensions[row].height = 52
        row += 1

    ov_idx = next((i for i, (f, _h) in enumerate(fields, start=1) if f == "overall"), None)
    if ov_idx and row > first:
        L = get_column_letter(ov_idx)
        r2 = f"{L}{first}:{L}{row - 1}"
        ws.conditional_formatting.add(r2, CellIsRule(
            operator="greaterThanOrEqual", formula=["2.5"], fill=FILL_HIGH, font=RATING_FONTS["High"]))
        ws.conditional_formatting.add(r2, CellIsRule(
            operator="between", formula=["1.5", "2.49"], fill=FILL_MED, font=RATING_FONTS["Medium"]))
        ws.conditional_formatting.add(r2, CellIsRule(
            operator="lessThan", formula=["1.5"], fill=FILL_LOW, font=RATING_FONTS["Low"]))

    ws.freeze_panes = f"A{first}"
    ws.auto_filter.ref = f"A{first - 1}:{get_column_letter(span)}{max(row - 1, first)}"
    ws.sheet_view.showGridLines = False


def training_summary(data):
    impacts = [i for i in data["impacts"] if i.get("training_required") == "Yes"]

    def _s(ws, row, span):
        by_method = defaultdict(lambda: [0, 0.0])
        by_group = defaultdict(lambda: [0, 0.0])
        for i in impacts:
            aud = i.get("training_audience_size") or i.get("headcount_impacted") or 0
            eff = (i.get("training_duration_hrs") or 0) * aud
            m = by_method[i.get("training_type") or "Unspecified"]
            m[0] += 1
            m[1] += eff
            g = by_group[i["stakeholder_group"]]
            g[0] += 1
            g[1] += eff

        for title, mapping, label in (("EFFORT BY DELIVERY METHOD", by_method, "Delivery method"),
                                      ("EFFORT BY AUDIENCE", by_group, "Stakeholder group")):
            section_header(ws, row, title, span)
            row += 1
            table_header(ws, row, [label, "Impacts", "Effort (person-hrs)", "Effort (days)"]
                         + [""] * (span - 4))
            row += 1
            first = row
            for k in sorted(mapping, key=lambda x: -mapping[x][1]):
                cnt, eff = mapping[k]
                for i, (v, fmt) in enumerate(
                        [(k, None), (cnt, "#,##0"), (eff, "#,##0"), (round(eff / 7.5, 1), "#,##0.0")],
                        start=1):
                    c = ws.cell(row=row, column=i, value=v)
                    c.border, c.font = CELL_BORDER, (BODY_BOLD if i == 1 else BODY)
                    c.alignment = Alignment(vertical="center", indent=1) if i == 1 else \
                        Alignment(horizontal="center", vertical="center")
                    if fmt:
                        c.number_format = fmt
                ws.row_dimensions[row].height = 18
                row += 1
            for i in range(1, 5):
                L = get_column_letter(i)
                v = "TOTAL" if i == 1 else f"=SUM({L}{first}:{L}{row - 1})"
                c = ws.cell(row=row, column=i, value=v)
                c.font, c.fill, c.border = BODY_BOLD, BAND, CELL_BORDER
                c.alignment = Alignment(indent=1) if i == 1 else Alignment(horizontal="center")
                c.number_format = "#,##0.0" if i == 4 else "#,##0"
            row += 2
        tip = ws.cell(row=row, column=1, value=(
            "Days assume 7.5 productive hours. This is business time away from the desk — the number "
            "to take to the sponsor when agreeing release of people for training. Check no group is "
            "carrying an unachievable load in the Readiness window (T-4 to T-1 weeks)."))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        tip.font, tip.alignment = SUBTLE, Alignment(wrap_text=True, vertical="center", indent=1)
        ws.row_dimensions[row].height = 28
        return row + 1
    return _s


def comms_summary(data):
    impacts = [i for i in data["impacts"] if i.get("comms_required") == "Yes"]

    def _s(ws, row, span):
        go_live = parse_date(data["meta"].get("go_live_date"))
        section_header(ws, row, "MESSAGES BY WAVE", span)
        row += 1
        table_header(ws, row, ["Wave", "Messages", "of which High", "Window", "Purpose"]
                     + [""] * (span - 5))
        row += 1
        for name, s_wk, e_wk, purpose in COMMS_WAVES:
            in_wave = [i for i in impacts if name.lower() in (i.get("comms_timing") or "").lower()]
            high = sum(1 for i in in_wave if computed_rating(i) == "High")
            window = (f"{go_live + timedelta(weeks=s_wk):%d %b} – {go_live + timedelta(weeks=e_wk):%d %b %Y}"
                      if go_live else f"T{s_wk:+d} to T{e_wk:+d} weeks")
            for i, v in enumerate([name, len(in_wave), high, window, purpose], start=1):
                c = ws.cell(row=row, column=i, value=v)
                c.border, c.font = CELL_BORDER, (BODY_BOLD if i == 1 else BODY)
                c.alignment = (Alignment(vertical="center", indent=1) if i in (1, 4, 5)
                               else Alignment(horizontal="center", vertical="center"))
                if i in (2, 3):
                    c.number_format = "#,##0"
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1
        tip = ws.cell(row=row, column=1, value=(
            "Sender matters more than channel: people accept organisational rationale from a senior "
            "sponsor and personal impact from their own line manager. High-rated impacts should appear "
            "in at least three waves — a single email in the Readiness window is not a comms plan."))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        tip.font, tip.alignment = SUBTLE, Alignment(wrap_text=True, vertical="center", indent=1)
        ws.row_dimensions[row].height = 28
        return row + 1
    return _s


def build_traceability(wb, data, last_live, extended):
    ws = wb.create_sheet("Traceability")
    meta = data["meta"]
    widths = [12, 24, 46, 14, 32, 14]
    set_widths(ws, widths)
    span = len(widths)

    style_title(ws, 1, "SOURCE TRACEABILITY", span)
    n = ws.cell(row=2, column=1, value=(
        "Every impact cites the documents it was derived from. When a business owner challenges a row "
        "in validation, this is what turns 'we disagree' into 'let's look at what the spec says'."))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    n.font, n.alignment = SUBTLE, Alignment(wrap_text=True, vertical="center", indent=1)

    counts = Counter()
    for imp in data["impacts"]:
        for r in split_refs(imp.get("source_ref")):
            counts[r] += 1

    row = 4
    section_header(ws, row, "SOURCE DOCUMENTS", span)
    row += 1
    table_header(ws, row, ["Ref", "Type", "Title", "Date", "Author / Participants", "Impacts derived"])
    row += 1
    for d in meta.get("source_documents", []):
        vals = [d.get("ref", ""), d.get("type", ""), d.get("title", ""),
                d.get("date", ""), d.get("author_or_participants", ""), counts.get(d.get("ref"), 0)]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = BODY_BOLD if i in (1, 6) else BODY
            c.border = CELL_BORDER
            c.alignment = (Alignment(horizontal="center", vertical="center") if i == 6
                           else Alignment(wrap_text=True, vertical="center", indent=1))
            if i == 6:
                c.number_format = "#,##0"
        ws.row_dimensions[row].height = 28
        row += 1
    ws.conditional_formatting.add(f"F6:F{row - 1}", CellIsRule(
        operator="equal", formula=["0"], fill=FILL_HIGH, font=RATING_FONTS["High"]))
    row += 2

    open_rows = [(DATA_START + i, imp) for i, imp in enumerate(data["impacts"])
                 if imp.get("confidence") == "Low" or (imp.get("notes") or "").strip()]
    section_header(ws, row, f"OPEN QUESTIONS FOR VALIDATION — {len(open_rows)} items", span)
    row += 1
    table_header(ws, row, ["Impact ID", "Stakeholder group", "Open question / note",
                           "Confidence", "L3 process", "Overall"])
    row += 1
    first = row
    for src_row, imp in open_rows:
        ws.cell(row=row, column=1, value=imp["impact_id"]).font = BODY_BOLD
        ws.cell(row=row, column=2, value=imp["stakeholder_group"]).font = BODY
        ws.cell(row=row, column=3, value=imp.get("notes", "")).font = BODY
        ws.cell(row=row, column=4, value=imp.get("confidence", "")).font = BODY
        ws.cell(row=row, column=5,
                value=lookup("l3", imp, f"$A{row}", last_live, extended, src_row)).font = BODY
        c = ws.cell(row=row, column=6,
                    value=lookup("overall", imp, f"$A{row}", last_live, extended, src_row))
        c.font, c.number_format = BODY_BOLD, "0.00"
        for i in range(1, span + 1):
            cc = ws.cell(row=row, column=i)
            cc.border = CELL_BORDER
            cc.alignment = (Alignment(horizontal="center", vertical="center") if i in (4, 6)
                            else Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[row].height = 46
        row += 1
    if row > first:
        ws.conditional_formatting.add(f"D{first}:D{row - 1}", CellIsRule(
            operator="equal", formula=['"Low"'], fill=FILL_HIGH, font=RATING_FONTS["High"]))
    ws.sheet_view.showGridLines = False


def build_assessment_info(wb, data, stats, extended):
    ws = wb.create_sheet("Assessment Info")
    meta = data["meta"]
    widths = [30, 26, 22, 22, 24, 22]
    set_widths(ws, widths)
    span = len(widths)

    style_title(ws, 1, "BASELINE CHANGE IMPACT ASSESSMENT", span, size=16)
    c = ws.cell(row=2, column=1, value=meta["program_name"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c.font = Font(name="Calibri", size=12, bold=True)
    c.alignment = Alignment(indent=1)

    row = 4
    section_header(ws, row, "ASSESSMENT DETAILS", span)
    row += 1
    for label, value in [
            ("Client", meta["client"]),
            ("Solution scope", meta["solution_scope"]),
            ("Assessment owner", meta["assessment_owner"]),
            ("Version", meta["version"]),
            ("Assessment date", meta["assessment_date"]),
            ("Target go-live", meta.get("go_live_date", "Not stated")),
            ("Impacts assessed", stats["total"]),
            ("Template", "Client CIA template — sheets 'CIA Template' and 'Change Impact Ratings' "
                         "carried through unchanged")]:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font, lc.fill, lc.border = BODY_BOLD, BAND, CELL_BORDER
        lc.alignment = Alignment(vertical="center", indent=1)
        vc = ws.cell(row=row, column=2, value=value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span)
        vc.font, vc.border = BODY, CELL_BORDER
        vc.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    section_header(ws, row, "IMPACT PROFILE", span)
    row += 1
    table_header(ws, row, ["Overall rating", "Impacts", "% of total", "Approx headcount",
                           "Training effort (person-hrs)", "Training effort (days)"])
    row += 1
    for label in RATINGS:
        d = stats["by_rating"][label]
        vals = [label, d["count"], (d["count"] / stats["total"]) if stats["total"] else 0,
                d["people"], d["effort"], round(d["effort"] / 7.5, 1)]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border, cell.font = CELL_BORDER, BODY
            cell.alignment = Alignment(horizontal="center" if i > 1 else "left", indent=0 if i > 1 else 1)
            cell.number_format = {3: "0%", 6: "#,##0.0"}.get(i, "#,##0")
        ws.cell(row=row, column=1).fill = RATING_FILLS[label]
        ws.cell(row=row, column=1).font = RATING_FONTS[label]
        row += 1
    for i, v in enumerate(["TOTAL", stats["total"], 1.0, stats["people_total"],
                           stats["effort_total"], round(stats["effort_total"] / 7.5, 1)], start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border, cell.font, cell.fill = CELL_BORDER, BODY_BOLD, BAND
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left", indent=0 if i > 1 else 1)
        cell.number_format = {3: "0%", 6: "#,##0.0"}.get(i, "#,##0")
    row += 2

    section_header(ws, row, "HOW THE RATING WORKS", span)
    row += 1
    for text, bold in [
        ("Each impact is scored 0-3 on People, Process and Technology against the anchors on the "
         "'Change Impact Ratings' sheet — the client's own rubric, carried through unchanged.", False),
        ("Overall Impact is the unweighted average of the three, as a live Excel formula. Re-score a "
         "Degree of Impact in a validation workshop and the average, heatmap and roll-ups all move.", False),
        ("ASSUMPTION — band cut-offs on the average: High ≥ 2.50 · Medium 1.50–2.49 · Low 0.50–1.49 · "
         "No/Minimal < 0.50. The template defines the per-dimension 0-3 scale but not the cut-offs for "
         "the overall average; these are this assessment's convention and should be confirmed with the "
         "client before baselining.", True),
        ("Anticipated resistance is tracked separately from impact magnitude — they are different "
         "things. A large welcome change is High impact / Low resistance; a small unwelcome one can be "
         "the reverse." + ("" if extended else " Recorded in the input data and shown on the Comms Plan "
                           "sheet; re-run with --extended to add it as a register column."), False),
        ("Confidence records how each row was derived: High = stated in a source document, Medium = "
         "inferred from two or more sources, Low = extrapolated. Low-confidence rows carry an open "
         "question and are listed on the Traceability sheet.", False),
    ]:
        cell = ws.cell(row=row, column=1, value="•  " + text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell.font = Font(name="Calibri", size=10, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[row].height = 42
        row += 1
    row += 1

    section_header(ws, row, "HOW TO USE THIS PACK", span)
    row += 1
    for text, bold in [
        ("CIA Template — the deliverable, in the client's own format. One row per process change × "
         "stakeholder group, filterable on the header row. Fifty blank pre-formatted rows with "
         "dropdowns and formulas sit under the data for adding impacts in-workbook.", False),
        ("Change Impact Ratings — the client's scoring rubric, untouched.", False),
        ("Impact Heatmap / Training Plan / Comms Plan / Traceability — supplementary views derived "
         "from the register. The template's Training and Communications columns hold the summary "
         "reference; the full detail lives on those sheets.", False),
        ("THIS IS A BASELINE, NOT A CONCLUSION. Low-confidence rows are inferred and carry an open "
         "question — work those with the business before baselining.", True),
    ]:
        cell = ws.cell(row=row, column=1, value="•  " + text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell.font = Font(name="Calibri", size=10, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[row].height = 40
        row += 1

    ws.sheet_view.showGridLines = False


def parse_date(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------------------
# Stats + orchestration
# --------------------------------------------------------------------------------------

def compute_stats(data):
    by_rating = {r: {"count": 0, "people": 0, "effort": 0.0} for r in RATINGS}
    people_total = effort_total = 0
    for imp in data["impacts"]:
        rating = computed_rating(imp) or "No / Minimal"
        hc = imp.get("headcount_impacted") or 0
        aud = imp.get("training_audience_size")
        aud = hc if aud in (None, "") else aud
        effort = (imp.get("training_duration_hrs") or 0) * (aud or 0)
        by_rating[rating]["count"] += 1
        by_rating[rating]["people"] += hc
        by_rating[rating]["effort"] += effort
        people_total += hc
        effort_total += effort
    return {"total": len(data["impacts"]), "by_rating": by_rating,
            "people_total": people_total, "effort_total": effort_total}


def build_workbook(data, out_path, template_path, extended):
    wb = load_workbook(template_path)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise SystemExit(f"Template has no '{TEMPLATE_SHEET}' sheet — got {wb.sheetnames}")

    stats = compute_stats(data)
    _last_data, last_live, _ov = fill_template_sheet(wb, data, extended)

    build_heatmap(wb, data, last_live, extended)
    build_rollup(
        wb, data, last_live, extended,
        sheet_name="Training Plan",
        title="TRAINING PLAN — derived from the impact register",
        blurb="Every impact flagged as needing training, with effort roll-ups by method and audience.",
        filter_field="training_required",
        fields=[("impact_id", "Impact ID"), ("stakeholder_group", "Audience"),
                ("l3", "L3 Process"), ("l4", "L4 Activity"), ("overall", "Overall"),
                ("training", "Training reference"), ("to_be", "What they must be able to do")],
        widths=[11, 26, 26, 28, 10, 42, 52],
        summary=training_summary(data))
    build_rollup(
        wb, data, last_live, extended,
        sheet_name="Comms Plan",
        title="COMMUNICATION PLAN — derived from the impact register",
        blurb="Key messages by audience and wave, with named senders.",
        filter_field="comms_required",
        fields=[("impact_id", "Impact ID"), ("stakeholder_group", "Audience"),
                ("overall", "Overall"), ("key_message", "Key Message"),
                ("benefit_narrative", "Benefit / WIIFM"), ("communications", "Comms reference"),
                ("resistance_risk", "Resistance"), ("mitigation_actions", "Mitigation")],
        widths=[11, 26, 10, 56, 42, 40, 12, 44],
        summary=comms_summary(data))
    build_traceability(wb, data, last_live, extended)
    build_assessment_info(wb, data, stats, extended)

    order = [TEMPLATE_SHEET, RUBRIC_SHEET, "Impact Heatmap", "Training Plan",
             "Comms Plan", "Traceability", "Assessment Info"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
    wb.active = 0
    wb.save(out_path)
    return stats


def main():
    ap = argparse.ArgumentParser(description="Populate the client CIA template from a JSON baseline.")
    ap.add_argument("input", help="Path to cia_input.json")
    ap.add_argument("-o", "--output", default="Change Impact Assessment.xlsx")
    ap.add_argument("-t", "--template", default=DEFAULT_TEMPLATE,
                    help="Client CIA template .xlsx (default: the vendored copy in templates/)")
    ap.add_argument("--extended", action="store_true",
                    help="Append governance columns (Impact ID, stakeholder group, resistance, "
                         "champion, source ref, confidence, status, notes) after the template's "
                         "columns. Off by default so the deliverable matches the client template.")
    ap.add_argument("--validate-only", action="store_true")
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as fh:
        data = json.load(fh)

    errors, warnings = validate(data)

    if errors:
        print(f"\n✗ {len(errors)} error(s) — workbook not generated:\n")
        for e in errors:
            print(f"  • {e}")
        print()
        return 1

    if warnings:
        print(f"\n⚠  {len(warnings)} warning(s) — review before baselining:\n")
        for w in warnings:
            print(f"  • {w}")
        print()
    else:
        print("\n✓ No validation warnings.\n")

    if args.validate_only:
        print("Validation only — no workbook written.")
        return 0

    stats = build_workbook(data, args.output, args.template, args.extended)
    print(f"✓ Wrote {args.output}"
          + ("  (template columns + governance extension)" if args.extended
             else "  (exact client template columns)"))
    print(f"  {stats['total']} impacts  |  " + "  ".join(
        f"{r}: {stats['by_rating'][r]['count']}" for r in RATINGS))
    print(f"  {stats['people_total']:,} headcount touchpoints  |  "
          f"{stats['effort_total']:,.0f} training person-hours "
          f"({stats['effort_total'] / 7.5:,.0f} days)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
